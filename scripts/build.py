#!/usr/bin/env python3
"""
build.py: consolidate the lexicon and regenerate every derived sheet.

ONE SOURCE OF TRUTH. Concept Roots holds every root. The Alphabetical Index,
Sound System, Phonotactics and Regeneration sheets are GENERATED from it and
from the converter, so they cannot drift out of date. Colours and Mod
Vocabulary keep their design rationale but stop being places a root is
defined.

Run this after any change to the rules or to Concept Roots.
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import root_converter as rc

# Paths. Override from the command line:
#   python3 build.py --in docs/lore/emerald_capitalism_language_system.xlsx --out same
# --out same writes back over the input, which is what you normally want.
SRC = "docs/lore/emerald_capitalism_language_system.xlsx"
OUT = "docs/lore/emerald_capitalism_language_system.xlsx"

HEAD = Font(name="Calibri", size=11, bold=True)
BODY = Font(name="Calibri", size=11)
GREEN = PatternFill("solid", fgColor="C6EFCE")
GREY = PatternFill("solid", fgColor="E7E6E6")
RED = PatternFill("solid", fgColor="FFC7CE")

COLS = ["Root", "Meaning", "Derivation", "Type", "Concept origin",
        "Shared concept", "Dialect forms", "Sporadic rules", "Stratum",
        "Former root", "Section", "Village naming"]

# The mod's village namer keys every root on its SECTION slug and reaches a
# section only through an axis. A root in a section no axis maps to can never
# be chosen, so the placeholder sections the consolidation created had to be
# replaced with real ones.
RESECTION = {
    "bonka": "Bank / exchange (institutional)", "pros": "Labor / making / transformation",
    "skona": "Knowledge / enchantment / ritual skill", "reska": "Knowledge / enchantment / ritual skill",
    "auka": "Bank / exchange (institutional)", "leikat": "Terrain / route / coast / natural placement",
    "die": "Labor / making / transformation", "stan": "Buildings / workstations / village function",
    "oova": "Knowledge / enchantment / ritual skill", "invent": "Quantity / sufficiency",
    "kapas": "Quantity / sufficiency",
    "enk": "Biome / material identity", "kook": "Biome / material identity",
    "lapa": "Biome / material identity", "popa": "Biome / material identity",
    "dandel": "Biome / material identity", "pek": "Biome / material identity",
    "lail": "Biome / material identity",
    "peo": "Biome / material identity", "oran": "Biome / material identity",
    "emhah": "Biome / material identity", "enboon": "Biome / material identity",
    "hahlapa": "Biome / material identity",
    "polapa": "Biome / material identity",
    "emchist": "Bank / exchange (institutional)", "emkool": "Protection / strength / communal defense",
    "empros": "Labor / making / transformation", "emdur": "Connection / transit",
    "temdur": "Connection / transit", "emsmet": "Labor / making / transformation",
    "viskraib": "Bank / exchange (institutional)", "vilbonka": "Bank / exchange (institutional)",
    "emhahdie": "Labor / making / transformation", "emhahstan": "Buildings / workstations / village function",
    "emhahbid": "Settlement / dwelling / social order",
    "koola": "Protection / strength / communal defense",
    "muknek": "Settlement / dwelling / social order",
}

# Roots that must never be a village name element. The mod concatenates two
# roots, so a specific manufactured item or a personal title reads as nonsense
# in a place name. Village named after an emerald green bed is not a village.
NO_VILLAGE_NAME = {
    # Manufactured items. The mod joins two roots, so a village named after an
    # emerald green bed is not a village.
    "emhahbid", "emhahstan", "emhahdie", "temdur", "emchist", "emdur",
    "stan", "die",
    # UI and interface terms. These name screens, not places.
    "invent", "kapas", "oova", "skona", "reska",
    # A personal title, and a Forced label with no derivation.
    "muknek", "poi",
    # wantreid names a PERSON, like muknek. A village called "wandering
    # trader" is the same category error as one called "mayor".
    "wantreid",
    # INSTITUTIONS, not features of a place. The bank section is otherwise
    # excellent place-name material: claims, tolls, standing, stored wealth,
    # which is the register real place names come from, but these six name
    # the institution itself. A village called "the village bank" is the same
    # category error as one called "emerald green bed".
    "vilbonka", "viskraib", "bonka", "auka", "inda", "frais",
}

PAD_VOWEL = {"armi": "i", "iglu": "u", "ingo": "o", "mansi": "i",
             "mopi": "i", "meiso": "o", "paki": "i"}
OVERRIDE = {
    "oran": ("HOMONYMY AVOIDANCE. Orange would give ora, but ora was already ore. "
             "The colour kept its nasal rather than collide. Compare orang, the "
             "stratum 4 profession title, which kept the whole syllable."),
    "orang": "S1 suspended. Stratum 4 froze almost unworn and this root kept its second syllable.",
    "muknek": "Stratum 1 substrate relic. Keeps a junction the language no longer allows.",
}

# Roots that live on the Colours and Mod Vocabulary sheets and have to move in.
# Compounds are built by rule C1 and are not derived from an English source.
INCOMING = [
    # (root, meaning, derivation, type, origin, sporadic, stratum)
    ("bonka", "held store of value, deposited wealth", "Bank", "Borrowed", "Village Bank", "S6", "3"),
    ("pros", "processing, conversion of one form to another", "Processor", "Borrowed", "Emerald Processor", "", "3"),
    ("skona", "sweep, survey, taking account of what is present", "Scan", "Borrowed", "Scan and rescan", "S6", "3"),
    ("reska", "repeated sweep, checking again", "Rescan", "Borrowed", "Scan and rescan", "S6", "3"),
    ("auka", "an account, a reckoning held in someone's name", "Account", "Borrowed", "Accounts", "", "3"),
    ("leikat", "a placed site, a fixed location", "Location", "Borrowed", "Golem Construction Location", "", "3"),
    ("die", "dye, applied colour", "Dye", "Borrowed", "Emerald Green Dye", "", "2b"),
    ("stan", "stained glass, coloured pane", "Stained Glass", "Borrowed", "Stained Glass", "", "2b"),
    ("oova", "overview, the whole seen at once", "Overview", "Borrowed", "Overview screen", "", "2b"),
    ("invent", "held goods, what is carried", "Inventory", "Borrowed", "Inventory", "", "2b"),
    ("kapas", "how much a thing can hold", "Capacity", "Borrowed", "Capacity", "", "2b"),
    ("enk", "black, and the ink it comes from", "Ink", "Borrowed", "Ink Sac / black dye", "", "2b"),
    ("kook", "brown, and the cocoa it comes from", "Cocoa", "Borrowed", "Cocoa Beans / brown dye", "", "2b"),
    ("lapa", "blue, and the lapis it comes from", "Lapis", "Borrowed", "Lapis Lazuli / blue dye", "S6", "2b"),
    ("popa", "red, and the poppy it comes from", "Poppy", "Borrowed", "Poppy / red dye", "S6", "2b"),
    ("dandel", "yellow, and the dandelion it comes from", "Dandelion", "Borrowed", "Dandelion / yellow dye", "", "2b"),
    ("pek", "lime, and the sea pickle it comes from", "Pickle", "Borrowed", "Sea Pickle / lime dye", "", "3"),
    ("lail", "magenta, and the lilac it comes from", "Lilac", "Borrowed", "Lilac / magenta dye", "", "3"),
    ("peo", "pink, and the peony it comes from", "Peony", "Borrowed", "Peony / pink dye", "", "3"),
    ("oran", "orange, the colour", "Orange", "Borrowed", "orange dye", "", "3"),
    # PROFESSION VOCABULARY. Only two new roots were needed; the rest of the
    # new professions are built by rule C1 from roots already here.
    ("wand", "wandering, moving without a fixed seat, travelling trade",
     "Wander", "Borrowed", "Wandering Trader", "", "3"),
    ("kot", "cutting, felling, separation by blade",
     "Cut", "Borrowed", "Axe / log chopping / Lumberjack", "", "3"),
]

INCOMING_COMPOUND = [
    ("emhah", "emerald green, the colour of emeralds", "emra + hah", "Compound", "Emerald Green", "", "5"),
    ("emchist", "emerald chest", "emra + chist", "Compound", "Emerald Chest", "", "5"),
    ("emkool", "emerald golem", "emra + kool", "Compound", "Emerald Golem", "", "5"),
    ("empros", "emerald processor", "emra + pros", "Compound", "Emerald Processor", "", "5"),
    ("emdur", "emerald door", "emra + dur", "Compound", "Emerald Door", "", "5"),
    ("temdur", "three-block emerald door", "te + emdur", "Compound", "Tall Emerald Door", "", "5"),
    ("emsmet", "emerald smith; as a title emsmeti", "emra + smet", "Compound", "Emerald Smith", "", "5"),
    ("viskraib", "village ledger", "vila + skraib", "Compound", "Village Ledger", "", "5"),
    ("vilbonka", "village bank", "vila + bonka", "Compound", "Village Bank", "", "5"),
    ("emhahdie", "emerald green dye", "emhah + die", "Compound", "Emerald Green Dye", "", "5"),
    ("emhahstan", "emerald green stained glass", "emhah + stan", "Compound", "Emerald Green Stained Glass", "", "5"),
    ("emhahbid", "emerald green bed", "emhah + bid", "Compound", "Emerald Green Bed", "", "5"),
    ("koola", "golem construction location", "kool + -a", "Compound", "Golem Construction Location", "", "5"),
    ("enboon", "grey", "enk + boon", "Compound", "grey dye", "", "5"),
    ("hahlapa", "cyan", "hah + lapa", "Compound", "cyan dye", "", "5"),
    ("polapa", "purple", "popa + lapa", "Compound", "purple dye", "", "5"),
    ("muknek", "mayor, an inherited title with no authority", "muk + nek", "Substrate", "pre-contact office", "", "1"),
    ("wantreid", "wandering trader, trade that comes to the village", "wand + treid", "Compound", "Wandering Trader", "", "5"),
    ("wokot", "wood-cutting, felling work, timber labour", "wot + kot", "Compound", "Lumberjack / Wood Cutter (mod)", "", "5"),
]

# The Root column above is what the rule PRODUCES. verify_compounds() rebuilds
# each one from its Derivation and fails the build on any mismatch, so a hand
# edit to a compound form cannot survive. LEXICALIZED forms are exempt: they
# predate rule C1 and are stored whole.


def verify_compounds():
    known = {row[0] for row in INCOMING_COMPOUND}
    bad = []
    for root, _meaning, parts, typ, _o, _s, strat in INCOMING_COMPOUND:
        if root in rc.LEXICALIZED or "+" not in parts:
            continue
        got = rc.build_from_parts(parts, known_compounds=known, stratum=strat)
        if got != root:
            bad.append((root, parts, got))
    if bad:
        for root, parts, got in bad:
            print(f"C1 MISMATCH: {parts} is recorded as {root} but the rule gives {got}")
        raise SystemExit("Compound verification failed. Fix the rule or mark the form LEXICALIZED.")
    return len(INCOMING_COMPOUND) - len(rc.LEXICALIZED)


def read_master(wb):
    """
    Read Concept Roots. Rows that share a root are a MERGE: three senses of
    Sculk became one word, so the senses join into one entry rather than
    sitting as three rows that look like duplicates.
    """
    ws = wb["Concept Roots"]
    entries, seen, by_root = [], set(), {}
    section = ""
    for row in ws.iter_rows():
        v = [c.value for c in row]
        a, b, c = (str(v[0] or "").strip(), str(v[1] or "").strip(), str(v[2] or "").strip())
        if a and not b and not c:
            section = a          # a row with only column A is a section header
            continue
        if v[0] in (None, "Root") or not v[3]:
            continue
        root = str(v[0]).strip()
        if root in seen:
            prev = by_root[root]
            extra = str(v[1] or "").strip()
            if extra and extra not in str(prev["meaning"]):
                prev["meaning"] = f"{prev['meaning']}; {extra}"
            for cell, key in ((v[4], "origin"), (v[6], "dialect")):
                if cell and str(cell) not in str(prev[key] or ""):
                    prev[key] = f"{prev[key]}; {cell}" if prev[key] else cell
            continue
        seen.add(root)
        entries.append({
            "root": root, "meaning": v[1], "derivation": v[2], "type": v[3],
            "origin": v[4], "shared": v[5], "dialect": v[6],
            "sporadic": v[7], "stratum": v[8],
            "former": v[9] if len(v) > 9 else None,
            "section": RESECTION.get(str(v[0]).strip(), section),
        })
        by_root[root] = entries[-1]
    return entries, seen


# Numerals live on the Grammar sheet and substrate elements on the Substrate
# sheet. Neither is a root, so a compound built from one is not dangling.
NON_ROOT_PARTS = {"un", "tu", "te", "fo", "faif", "sik", "sef", "et", "ploh",
                  "stak", "muk", "nek", "-a", "-i", "-o", "-ek", "-ur", "-in",
                  "-na"}


def refresh_compound_parts(wb, entries):
    """
    A compound's Derivation cell names its parts by root. When a root changed,
    those cells kept pointing at the old form: kovolt still said cov + valt.
    Rewrite them from the Former root column so a compound always names roots
    that exist.
    """
    former = {}
    for e in entries:
        if e["former"]:
            former[str(e["former"]).strip()] = e["root"]
    roots = {e["root"] for e in entries}
    fixed = []
    for sheetname in ("Concept Roots",):
        for row in wb[sheetname].iter_rows():
            v = [c.value for c in row]
            if v[0] in (None, "Root") or not v[3]:
                continue
            src = str(v[2] or "")
            if "+" not in src:
                continue
            parts = [p.strip() for p in src.split("+")]
            new = [former.get(p, p) if p not in roots else p for p in parts]
            if new != parts:
                row[2].value = " + ".join(new)
                row[2].font = BODY
                fixed.append((str(v[0]).strip(), src, " + ".join(new)))
                for e in entries:
                    if e["root"] == str(v[0]).strip():
                        e["derivation"] = " + ".join(new)
    return fixed


def add_incoming(entries, seen, ws):
    added = 0
    for label, rows in (("MOD AND COLOUR VOCABULARY, borrowed", INCOMING),
                        ("BUILT BY RULE, compounds and substrate", INCOMING_COMPOUND)):
        ws.append([None] * len(COLS))
        c = ws.cell(row=ws.max_row, column=1, value=label)
        c.font = HEAD
        ws.append(COLS)
        for cell in ws[ws.max_row]:
            cell.font = HEAD
        for root, meaning, deriv, typ, origin, spor, strat in rows:
            if root in seen:
                continue
            seen.add(root)
            ws.append([root, meaning, deriv, typ, origin, None, None,
                       spor or None, strat, None])
            for cell in ws[ws.max_row]:
                cell.font = BODY
            ws.cell(row=ws.max_row, column=1).fill = GREEN
            entries.append({"root": root, "meaning": meaning, "derivation": deriv,
                            "type": typ, "origin": origin, "shared": None,
                            "dialect": None, "sporadic": spor or None,
                            "stratum": strat, "former": None})
            added += 1
    return added


def verdict(e):
    """Does the converter still produce this root?"""
    root, typ = e["root"], str(e["type"] or "")
    src = str(e["derivation"] or "").strip()
    if root in OVERRIDE:
        return "lexical override", OVERRIDE[root], ""
    if typ.lower() in ("compound", "substrate") or "+" in src:
        return "built by rule", "compound rule C1, not derived from English", ""
    if rc.is_forced(typ) or typ.lower() != "borrowed" or not src or " " in src:
        return "not derived", "forced, or no single-word source", ""
    spor = str(e["sporadic"] or "")
    kw = {"aphaeresis": "S5" in spor, "paragoge": "S6" in spor,
          "breaking": "S7" in spor,
          # S9 records WHICH site fired, as S9(1). Bare S9 means site 0.
          "metathesis": rc.metathesis_site(spor) if "S9" in spor else False}
    if root in PAD_VOWEL:
        kw["paragoge"] = True
    got, ok, why = rc.convert(src, rc.norm_stratum(e["stratum"]), verbose=False,
                              pad_vowel=PAD_VOWEL.get(root), **kw)
    if got == root:
        note = f"lexical pad vowel -{PAD_VOWEL[root]}" if root in PAD_VOWEL else ""
        return "reproduces", note, got
    return "FAILS", "the rules no longer produce this root", got


def sheet(wb, name, position):
    if name in wb.sheetnames:
        del wb[name]
    return wb.create_sheet(name, position)


def write_lines(s, lines, start=1):
    r = start
    for text, bold in lines:
        c = s.cell(row=r, column=1, value=text)
        c.font = HEAD if bold else BODY
        r += 1
    return r


# ---------------------------------------------------------------- DIALECTS
# The Dialects sheet states the rule itself: forms are GENERATED, never
# authored, and regenerated whenever a root changes. One rule per region,
# applied to the standard form, then constraint-checked. A variant that
# breaks the phonotactics is discarded and the standard form stands.
# Compounds take the rule on the JOINED form, not on the parts.

def _shift_short(word, frm, to):
    """Raise a SHORT vowel. A long nucleus is a different segment and does not
    raise: ei and ii are not e and i, and oo is not o."""
    segs = rc.tokenize(word)
    return rc.untokenize([to if s == frm else s for s in segs])


def _apocope(word):
    """Delete a final consonant SEGMENT. th and ng go as a unit."""
    segs = rc.tokenize(word)
    if len(segs) > 1 and not rc.is_vowel(segs[-1]):
        return rc.untokenize(segs[:-1])
    return word


LENITION = {"p": "b", "t": "d", "k": "g",   # voiceless stop voices
            "v": "w"}                       # voiced fricative becomes a glide


def _lenite(word):
    """
    Savanna's second rule. A consonant between two vowels weakens one step on
    the lenition scale: a voiceless stop voices, p > b, t > d, k > g, and a
    voiced fricative becomes a glide, v > w. Degemination alone generated ONE
    form in 215, because the derivation already degeminates and almost nothing
    reaches the dialect stage with a geminate left. This gives Savanna a rule
    that actually fires, and it keeps the region's identity consonantal while
    Desert and Taiga move vowels and Snowy deletes.

    The v > w step is the same process one notch further along, not a second
    rule. It exists because a root ending in a vowel with no /e/ and no /o/ was
    invisible to every region: rava had no dialect form anywhere.
    """
    segs = rc.tokenize(word)
    out = list(segs)
    for i in range(1, len(segs) - 1):
        if segs[i] in LENITION and rc.is_vowel(segs[i - 1]) and rc.is_vowel(segs[i + 1]):
            out[i] = LENITION[segs[i]]
    return rc.untokenize(out)


def _savanna(word):
    return _lenite(rc.degeminate(word))


REGIONS = [
    ("Des", lambda w: _shift_short(w, "e", "i")),   # Desert, /e/ > /i/
    ("Sav", _savanna),                               # Savanna, CC > C and intervocalic voicing
    ("Tai", lambda w: _shift_short(w, "o", "u")),   # Taiga, /o/ > /u/
    ("Sno", _apocope),                               # Snowy, apocope
]                                                    # Plains is the standard


def build_dialects(entries, wb):
    stats = {tag: 0 for tag, _ in REGIONS}
    blocked = {tag: 0 for tag, _ in REGIONS}
    homonym = []
    standard = {e["root"] for e in entries}
    forms = {}
    for e in entries:
        std = e["root"]
        strat = rc.norm_stratum(e["stratum"])
        comp = (str(e["type"] or "").lower() in ("compound", "substrate")
                or "+" in str(e["derivation"] or ""))
        out = []
        for tag, rule in REGIONS:
            got = rule(std)
            if got == std:
                continue
            ok, _ = rc.legal(got, strat, compound=comp)
            if not ok:
                blocked[tag] += 1
                continue
            if got in standard and got != std:
                # The variant would be a word that already means something
                # else. A speaker in that region would have no way to tell
                # them apart, so the rule is blocked and the standard stands.
                homonym.append((tag, std, got))
                blocked[tag] += 1
                continue
            out.append(f"{tag} {got}")
            stats[tag] += 1
        forms[std] = "; ".join(out) if out else None
        e["dialect"] = forms[std]
    for row in wb["Concept Roots"].iter_rows():
        v = [c.value for c in row]
        if v[0] in (None, "Root") or not v[3]:
            continue
        root = str(v[0]).strip()
        if root in forms:
            row[6].value = forms[root]
            row[6].font = BODY
    return stats, blocked, homonym


# ------------------------------------------------------------------ NAMES
# The SUBSTRATE NAMES sheet is not regenerated here and must not be. It holds
# 1,039 BASE forms built from the substrate element pairs, a closed set of 20
# plus 20 that no root change can touch. It is not a shipped name list: every
# village applies its own drift rules on top, so the form a player sees is
# computed at worldgen. See the DRIFT rows on the Substrate sheet.
#
# The Substrate sheet is hand-edited and is not regenerated here either.
#
# NAMES AND PROFESSIONS is the sheet that goes stale, because a profession
# byname is a ROOT plus -i, and nine of the fourteen roots have moved.

PROFESSIONS = [
    ("Armorer", "armi"), ("Butcher", "smoka"), ("Cartographer", "mapi"),
    ("Cleric", "boti"), ("Farmer", "homfa"), ("Fisherman", "roda"),
    ("Fletcher", "letch"), ("Leatherworker", "kalda"), ("Librarian", "lekt"),
    ("Mason", "maso"), ("Shepherd", "wlu"), ("Toolsmith", "pika"),
    ("Weaponsmith", "wepa"),
    # Vanilla, newly given a root.
    ("Wandering Trader", "wantreid"),
    # Mod professions.
    ("Banker (mod)", "enda"),
    ("Emerald Smith (mod)", "emsmet"),
    ("Merchant (mod)", "treid"),
    ("Lumberjack (mod)", "wokot"),
]

# Professions that take NO byname, with the reason. Mayor is not an oversight:
# muknek is a defective noun that refuses every affix.
NO_BYNAME = [
    ("Mayor (mod)", "NO BYNAME. muknek is a defective noun: no affixes, and never "
                    "the -i agent suffix. Stratum 1 relic. Do not 'fix' this."),
    ("Nitwit", "No byname. Keeps -ek, the CHILD suffix, for life. DELIBERATE: the "
               "village never regards a nitwit as having grown up. Do not 'fix' this to -in."),
    ("Unemployed", "No byname at all."),
]

REDUNDANCY = [
    # sela settlement and vena extent were named here but never existed as
    # roots anywhere in the lexicon. Removed rather than invented.
    ("Dwelling", [("hala", "house"), ("gard", "interior"), ("vila", "kin"),
                  ("beid", "bed"), ("beir", "headcount")]),
    ("Water edge", [("mera", "still"), ("rana", "running"), ("resa", "distance"),
                    ("rida", "reeds"), ("rim", "boundary"), ("brid", "crossing")]),
    ("Metalwork", [("mel", "everyday"), ("anva", "striking"), ("furna", "smelting"),
                   ("ingo", "product"), ("grid", "correction")]),
]


def byname(root):
    """
    Profession byname: the root plus -i, under RULE A1. n is INSERTED before a
    vowel-initial suffix on a vowel-final stem. Nothing is deleted, so the old
    only-nucleus guard is unnecessary and has been removed.

    roda gives Rodani, not Rodi. kold is unchanged at Koldi. The Snowy Cleric
    gives Buni rather than the hiatus form Bui.
    """
    return rc.affix(root, "-i")[0].capitalize()


def build_names(wb, entries):
    current = {e["root"] for e in entries}
    former = {str(e["former"]).strip(): e["root"] for e in entries if e["former"]}

    def now(old):
        return old if old in current else former.get(old)

    s = sheet(wb, "Names and Professions", 8)
    r = write_lines(s, [
        ("NAME GENERATOR AND PROFESSIONS. GENERATED by build.py from the current roots.", True),
        ("", False),
        ("A byname is the profession root plus -i. Rule A1 inserts n after a vowel-final stem; nothing is deleted.", False),
        ("The pending-shift notes are gone: the shift decisions are settled and these are the derived forms.", False),
        ("Substrate Names is not regenerated here. Those are element pairs, not roots.", False),
        ("It holds BASE forms. Each village applies its own drift rules on top, so the shipped form is computed at worldgen.", False),
        ("", False),
    ])
    for j, h in enumerate(["Type", "Item", "Root", "Byname (Plains)",
                           "Byname Des", "Byname Sav", "Byname Tai",
                           "Byname Sno", "Note"], start=1):
        s.cell(row=r, column=j, value=h).font = HEAD
    r += 1
    missing = []
    dia_by_root = {e["root"]: (e.get("dialect") or "") for e in entries}
    for prof, old in PROFESSIONS:
        root = now(old)
        if root is None:
            missing.append((prof, old))
            vals = ("PROFESSION", prof, old, "", "", "", "", "",
                    "ROOT MISSING from the lexicon")
        else:
            # A byname is built from the form the REGION uses, not from the
            # standard. A desert Fletcher is a Flichi, because the desert root
            # is flich. Village drift then applies on top of that.
            forms = {}
            for part in (dia_by_root.get(root) or "").split(";"):
                part = part.strip()
                if " " in part:
                    tag, form = part.split(" ", 1)
                    forms[tag] = form.strip()
            note = f"was {old}" if root != old else ""
            vals = ("PROFESSION", prof, root, byname(root),
                    byname(forms["Des"]) if "Des" in forms else "",
                    byname(forms["Sav"]) if "Sav" in forms else "",
                    byname(forms["Tai"]) if "Tai" in forms else "",
                    byname(forms["Sno"]) if "Sno" in forms else "",
                    note)
        for j, v in enumerate(vals, start=1):
            s.cell(row=r, column=j, value=v).font = BODY
        if root and root != old:
            s.cell(row=r, column=3).fill = GREEN
        if root is None:
            s.cell(row=r, column=3).fill = RED
        r += 1
    for item, note in NO_BYNAME:
        for j, v in enumerate(("PROFESSION", item, "", "", "", "", "", "", note), start=1):
            s.cell(row=r, column=j, value=v).font = BODY
        r += 1

    r += 1
    for j, h in enumerate(["Slot", "Item", "Value"], start=1):
        s.cell(row=r, column=j, value=h).font = HEAD
    r += 1
    for slot, item, value in [
        ("NAME SLOT 1", "Personal root", "Substrate element pair. The biome pool decides which elements. Fixed at spawn."),
        ("NAME SLOT 2", "Personal suffix", "DETERMINED BY AGE, never drawn at random: -ek child, -in adult, -ur elder. It CHANGES as the villager ages."),
        ("NAME SLOT 2", "Expected mix", "About 12pct -ek, 80pct -in, 8pct -ur in a settled village. That is the RESULT of the age spread, not a weighting to sample from."),
        ("NAME SLOT 3", "Profession byname", "Profession root plus -i, built from the REGION's form of the root, not the standard. Changes with profession. Omit for unemployed."),
        ("NAME SLOT 4", "Origin particle", "Se plus the origin village name AS COINED. Omit if never migrated."),
        ("DRIFT SCOPE", "Slots 1, 2 and 3", "The village's drift rules apply to slots 1, 2 and 3. A sound change is blind to etymology, so it reaches the suffix and the byname too."),
        ("DRIFT SCOPE", "Slot 4 exempt", "The origin particle does NOT drift. It carries the origin village's name as coined. Village names fossilise and a migrant does not rename their birthplace."),
        ("DRIFT SCOPE", "Order", "Region first, then village. Resolve the biome dialect form, build the byname, then apply the village's drift rules in ascending rule index (D1 before D3, D3 before D9)."),
        ("DRIFT SCOPE", "Safety", "Checked: no two professions collide under any of the 77 rule combinations, and -ek, -in and -ur stay distinct under every rule. 77, not 78: 12 singles plus C(12,2)=66 pairs minus the D9/D10 inverse, which never co-occur. Re-verified under rule A1 for all 17 byname-bearing villager professions, 385 cases, 0 clashes."),
        ("NAME FORM", "Full", "Bemmunin Mopi sa-Bemmun"),
        ("NAME FORM", "Dialogue", "Bemmunin"),
    ]:
        for j, v in enumerate((slot, item, value), start=1):
            s.cell(row=r, column=j, value=v).font = BODY
        r += 1

    r += 1
    s.cell(row=r, column=1, value="REDUNDANCY SETS, refreshed to the current roots").font = HEAD
    r += 1
    gone = []
    for label, items in REDUNDANCY:
        parts = []
        for old, sense in items:
            root = now(old)
            if root is None:
                gone.append((label, old, sense))
                parts.append(f"[{old} MISSING] {sense}")
            else:
                parts.append(f"{root} {sense}" + (f" (was {old})" if root != old else ""))
        for j, v in enumerate(("REDUNDANCY", label, ", ".join(parts)), start=1):
            c = s.cell(row=r, column=j, value=v)
            c.font = BODY
            c.alignment = Alignment(wrap_text=(j == 3), vertical="top")
        r += 1
    for col, w in zip("ABCDE", (16, 24, 14, 14, 104)):
        s.column_dimensions[col].width = w
    return missing, gone


def build_strata(wb):
    """
    The Strata sheet listed retired rules (S2, S4) and a stratum 4 pipeline of
    S0 S1 S9. It also said stratum 2 takes no S6 and is consonant-final by
    design, which the lexicon contradicts: 12 of 43 stratum 2a roots end in -a.
    Rewrite the rule columns from STRATA and the pipelines.
    """
    g = wb["Strata"]
    seen, drop = set(), []
    for r in g.iter_rows():
        key = tuple(str(c.value) for c in r[:5])
        if key in seen and r[0].value:
            drop.append(r[0].row)
        seen.add(key)
    for n in reversed(drop):
        g.delete_rows(n)
    pipe = {"2": "S0b S0 S1 | S5* S8 S7* S6* P9 S3 S3b S3c P10 S9*  (2a; 2b drops S8)",
            "2a": "S0b S0 S1 | S5* S8 S7* S6* P9 S3 S3b S3c P10 S9*",
            "2b": "S0b S0 S1 | S5* S7* S6* P9 S3 S3b S3c P10 S9*  (no S8)",
            "3":  "S0b S0 S1 | S5* S7* S6* P9 S3 S3b S3c P10 S9*  (no S8)",
            "4":  "S0b S0 S1 (suspended at two syllables or fewer) | S5* S11 S6* S9*",
            "5":  "NONE. Built from existing roots by compound rule C1."}
    fixed = 0
    for row in g.iter_rows():
        key = str(row[0].value or "").strip()
        if key in pipe and row[3].value:
            row[3].value = pipe[key]
            row[3].font = BODY
            fixed += 1
        if key == "Which rules fire":
            st = str(row[1].value or "").strip()
            if st in pipe:
                row[2].value = pipe[st]
                row[2].font = BODY
                fixed += 1
            row[3].value = ("Rules marked * are SPORADIC and fire only when the root records them: "
                            "S5, S6, S7, S9.")
            row[3].font = BODY
        if key == "Sporadic rules":
            row[2].value = "S5 aphaeresis, S6 paragoge, S7 breaking, S9 metathesis"
            row[3].value = ("Never automatic at any stratum. S6 fires on 25 percent of borrowed roots and has no "
                            "phonological conditioning, so stratum 2 is NOT consonant-final by design.")
            for c in row:
                c.font = BODY
            fixed += 1
    return fixed


def build_dialect_examples(wb, entries):
    """The Dialects sheet illustrated each rule with roots that no longer
    exist: mera > mira, bela > bila, ward > war, pilost > pilust. Pull live
    examples from the forms the build just generated."""
    by_tag = {}
    for e in entries:
        for part in str(e["dialect"] or "").split("; "):
            if not part:
                continue
            tag, form = part.split(" ", 1)
            by_tag.setdefault(tag, []).append(f"{e['root']} > {form}")
    for row in wb["Dialects"].iter_rows():
        if isinstance(row[3].value, str) and "pilost" in row[3].value:
            row[3].value = ("Apply the rule to the joined form: pilpoost gives Taiga pilpuust, "
                            "not pil plus puust.")
            row[3].font = BODY
    label = {"Desert": "Des", "Savanna": "Sav", "Taiga": "Tai", "Snowy": "Sno"}
    fixed = 0
    for row in wb["Dialects"].iter_rows():
        region = str(row[0].value or "").strip()
        tag = label.get(region)
        if tag and row[3].value:
            ex = ", ".join(by_tag.get(tag, [])[:3])
            note = {"Desert": "", "Savanna": "Degemination plus intervocalic voicing. ",
                    "Taiga": "", "Snowy": "BLOCKED where the result is illegal or lands on another root. "}[region]
            row[3].value = note + (ex or "no forms generated")
            row[3].font = BODY
            fixed += 1
    return fixed


def build_index(wb, entries):
    s = sheet(wb, "Alphabetical Index", 5)
    r = write_lines(s, [
        ("ALPHABETICAL INDEX. GENERATED from Concept Roots by build.py. Do not edit here; edit Concept Roots and rebuild.", True),
        ("", False),
    ])
    for j, h in enumerate(COLS, start=1):
        s.cell(row=r, column=j, value=h).font = HEAD
    r += 1
    for e in sorted(entries, key=lambda x: x["root"]):
        vals = [e["root"], e["meaning"], e["derivation"], e["type"], e["origin"],
                e["shared"], e["dialect"], e["sporadic"], e["stratum"], e["former"],
                e["section"], "no" if e["root"] in NO_VILLAGE_NAME else "yes"]
        for j, v in enumerate(vals, start=1):
            s.cell(row=r, column=j, value=v).font = BODY
        r += 1
    for col, w in zip("ABCDEFGHIJKL", (12, 56, 16, 12, 30, 18, 20, 14, 9, 12, 34, 13)):
        s.column_dimensions[col].width = w


def build_sound_system(wb):
    s = sheet(wb, "Sound System", 10)
    r = write_lines(s, [
        ("SOUND SYSTEM. GENERATED from root_converter.py by build.py. Do not edit here; edit the converter and rebuild.", True),
        ("", False),
        ("The stratum decides which rules fire. Rules marked SPORADIC never fire on their own: a root records that it took", False),
        ("one, and the converter applies it only then.", False),
        ("", False),
    ])
    for j, h in enumerate(["Order", "Rule", "Name", "Category", "Applies to", "Statement"], start=1):
        s.cell(row=r, column=j, value=h).font = HEAD
    r += 1
    rows = [
        ("1", "G2P", "source pronunciation", "preprocessing", "all",
         "English spelling to one sound per segment. A lookup table handles words the heuristic gets wrong, then soft c, silent e, digraphs, degemination."),
        ("2", "S0b", "vowel length", "historical", "all",
         "A tense English vowel becomes a two-segment nucleus: ei, ih, ai, oo, uu, au. Long /i:/ is written ii instead of ih where an h is already in the word, or where the shift will create one."),
        ("3", "S0", "stress", "preprocessing", "all",
         "English primary stress. Latinate shapes stress before -tion, -ity, -ic and friends, and truncate THROUGH the stressed syllable rather than starting at it."),
        ("4", "S1", "truncation", "historical", "all",
         "Cut to the stressed syllable. Syllabification is coda-heavy, so the stressed syllable already carries the consonant after its vowel. Golem parses gol-em and gives gol."),
        ("5", "S5", "aphaeresis", "SPORADIC", "all",
         "Delete a single word-initial consonant. Three attestations only."),
        ("6", "S8", "the shift", "historical", "2a only",
         "p>f, t>th, k>h, b>p, d>t, g>k, v>f. A chain, not a swap. Runs after S5 so losses precede shifts."),
        ("7", "S7", "breaking", "SPORADIC", "all",
         "i > ai before a resonant. Runs before any padding, because it changes the stem vowel."),
        ("8", "S6", "paragoge", "SPORADIC", "all",
         "Add a final -a. Fires on 25 percent of borrowed roots and has no phonological conditioning. Seven roots take a different vowel, recorded on the root."),
        ("9", "P9", "minimal word", "repair", "2a 2b 3",
         "A root of fewer than three SEGMENTS takes the paragogic vowel. Counted in segments, so ch and a long nucleus each count once."),
        ("10", "S3", "cluster reduction", "historical", "all",
         "An illegal final cluster trims from the right one segment at a time until it is legal. Never below three segments; see P10."),
        ("11", "S3b", "monosyllabic coda", "historical", "2a 2b 3",
         "A LEGAL final cluster survives. What remains of this rule is th-stopping: Path gives pot, Smith gives smet."),
        ("12", "S3c", "compensatory raising", "historical", "2a 2b 3",
         "A one-syllable source raises its stem vowel one step: a>o, e>i, o>u, u>o, i>e. This is the whole of monosyllabic wear. A long nucleus does not raise."),
        ("13", "P10", "minimum content", "repair", "2a 2b 3",
         "No reduction leaves fewer than three segments. A cluster S3 was forbidden to cut takes the paragogic vowel instead. Emerald gives emra, not em."),
        ("14", "S11", "final -er to -a", "historical", "4 only",
         "Stratum 4's one change of its own. Unstressed final -er reduces to -a: Nether gives netha. Guarded on "
         "syllable count, so a monosyllable whose only nucleus is that e is untouched and per stays per."),
        ("15", "S9", "metathesis", "SPORADIC", "all",
         "Transpose a liquid or glide with an adjacent vowel. A word can have several eligible sites and which one moves is a property of the word."),
        ("16", "C1", "compounding", "word formation", "5",
         "Head-final: modifier first, head last, only the final element takes a suffix. Over two syllables a SIMPLE modifier clips to its first syllable. A compound modifier does not clip again."),
    ]
    for row in rows:
        for j, v in enumerate(row, start=1):
            c = s.cell(row=r, column=j, value=v)
            c.font = BODY
            c.alignment = Alignment(wrap_text=(j == 6), vertical="top")
        if row[3] == "SPORADIC":
            s.cell(row=r, column=4).fill = GREY
        r += 1
    r += 1
    write_lines(s, [
        ("RETIRED. S2 syncope and S4 final stop loss each fired on one root and live in retired_rules.py.", True),
        ("The one case S2 earned is folded into S1 as bounded syncope: Emerald gives emr, then P10 repairs it to emra.", False),
    ], r)
    for col, w in zip("ABCDEF", (7, 8, 22, 14, 12, 104)):
        s.column_dimensions[col].width = w


def build_phonotactics(wb, entries):
    s = sheet(wb, "Phonotactics", 13)
    r = write_lines(s, [
        ("PHONOTACTICS. GENERATED from root_converter.py by build.py. These say when an OUTPUT is illegal.", True),
        ("", False),
    ])
    for j, h in enumerate(["Position", "Constraint", "Detail"], start=1):
        s.cell(row=r, column=j, value=h).font = HEAD
    r += 1
    rows = [
        ("Segments", "One letter is not one sound",
         "Single segments: " + ", ".join(rc.DIGRAPHS) + ". Everything counts these once."),
        ("Vowels", "a e i o u, plus the long nuclei",
         "Long: " + ", ".join(rc.LONG_NUCLEI) + ". A long nucleus is ONE segment and ONE syllable."),
        ("Consonants", "b d f g h j k l m n p r s t v w",
         "c, q, x and y are not segments at all. z is legal at strata 4 and 5 only, where the vocabulary froze before the phonology absorbed it."),
        ("Whole root", "A root must have a vowel, and cannot be one segment repeated",
         "Guarded on length, since a single long nucleus is one segment: ai is a word, aa is not."),
        ("Whole root", f"At most {rc.MAX_NUCLEI} syllables",
         "SIMPLE roots in strata 2 and 3 only. Compounds are exempt, since each half already satisfies it. Strata 4 and 5 are exempt, having frozen before the wear that shortens roots."),
        ("Whole root", f"At least {rc.MIN_CONTENT} segments",
         "Enforced by P9 as a repair and by P10 as a limit on reduction."),
        ("Onset", "Sonority, not a list",
         "Obstruent plus liquid or glide, or s plus a consonant, or w plus a liquid. Plus the digraphs, which are single segments. Maximum three segments."),
        ("Coda", "Two segments, from the legal set",
         ", ".join(sorted(rc.LEGAL_CODA_2))),
        ("Coda", "Never three or more segments", "S3 trims from the right until legal, subject to P10."),
        ("Strata 4 and 5", "Exempt from the cluster rules and the syllable ceiling",
         "They froze almost unworn. Distinctness and the vowel requirement still apply."),
    ]
    for row in rows:
        for j, v in enumerate(row, start=1):
            c = s.cell(row=r, column=j, value=v)
            c.font = BODY
            c.alignment = Alignment(wrap_text=(j == 3), vertical="top")
        r += 1

    bad = []
    for e in entries:
        comp = str(e["type"] or "").lower() in ("compound", "substrate") or "+" in str(e["derivation"] or "")
        ok, why = rc.legal(e["root"], rc.norm_stratum(e["stratum"]), compound=comp)
        if not ok:
            bad.append((e["root"], str(e["stratum"]), "; ".join(why)))
    r += 1
    s.cell(row=r, column=1, value=f"CURRENT VIOLATIONS: {len(bad)} of {len(entries)} roots").font = HEAD
    r += 1
    for row in bad:
        for j, v in enumerate(row, start=1):
            s.cell(row=r, column=j, value=v).font = BODY
        s.cell(row=r, column=1).fill = RED
        r += 1
    for col, w in zip("ABC", (18, 46, 104)):
        s.column_dimensions[col].width = w


def build_regeneration(wb, entries):
    s = sheet(wb, "Build Report", 3)
    counts = {"reproduces": 0, "lexical override": 0, "built by rule": 0,
              "not derived": 0, "FAILS": 0}
    rows = []
    for e in entries:
        v, note, got = verdict(e)
        counts[v] += 1
        rows.append((e["root"], e["derivation"], e["stratum"], e["sporadic"],
                     got, v, note, str(e["meaning"] or "")[:56]))
    r = write_lines(s, [
        ("REGENERATION. GENERATED by build.py. This is a REGRESSION HARNESS, not a fitting report.", True),
        ("", False),
        ("Every borrowed root is the converter's own output, so the question is no longer how much of the lexicon the rules", False),
        ("can reproduce. It is whether they still do. A row marked FAILS means a rule changed and the lexicon was not rebuilt.", False),
        ("", False),
        (f"  reproduces from source, stratum and sporadic rules   {counts['reproduces']}", False),
        (f"  lexical override, a claim about one word              {counts['lexical override']}", False),
        (f"  built by rule C1, not derived                         {counts['built by rule']}", False),
        (f"  not derived, forced or multi-word source              {counts['not derived']}", False),
        (f"  FAILS                                                 {counts['FAILS']}", False),
        ("", False),
    ])
    for j, h in enumerate(["Root", "Derivation", "Stratum", "Sporadic",
                           "Converter output", "Verdict", "Note", "Meaning"], start=1):
        s.cell(row=r, column=j, value=h).font = HEAD
    r += 1
    for row in sorted(rows):
        for j, v in enumerate(row, start=1):
            s.cell(row=r, column=j, value=v).font = BODY
        c = s.cell(row=r, column=6)
        c.fill = {"reproduces": GREEN, "FAILS": RED}.get(row[5], GREY)
        r += 1
    for col, w in zip("ABCDEFGH", (14, 16, 10, 12, 16, 18, 62, 56)):
        s.column_dimensions[col].width = w
    return counts, s, r


KNOWN_TYPES = {"borrowed", "compound", "forced", "substrate"}

OPEN_DECISIONS = [
    ("Stratum 4 pass-through", "watch",
     "DECIDED: S1 is suspended at stratum 4 for a source of two syllables or fewer, so the layer arrives almost "
     "unworn. That is what stratum 4 means, and it gives orang without an override. The COST is that 73 of 300 "
     "unseen words now pass through stratum 4 unchanged, up from 50, brought back to 70 by S11, and roots like warden and trial are "
     "identical to their English sources. Strata 2a, 2b and 3 remain at zero. Reversible: set s1_keep to 0."),
    ("Native word formation", "open",
     "Verbs, and productive derivation beyond -a, -i, -o, -ek, -ur and the new possessive -na. Deliberately minimal, "
     "since villagers cannot express relations between nouns."),
]


def build_adopt(wb, entries, s, r):
    """
    GENERATED. Was a one-off list of roots with no derivation, all long since
    resolved. It is now a live issues sheet: problems the build DETECTS in the
    data, plus decisions that are recorded as open.
    """
    roots = {e["root"] for e in entries}
    found = []
    for e in entries:
        typ = str(e["type"] or "").strip()
        if typ.lower() not in KNOWN_TYPES:
            found.append(("data error", e["root"],
                          f"Type column reads '{typ}', which is not a known type. The row is skipped by every "
                          f"generated sheet until this is fixed."))
        src = str(e["derivation"] or "").strip()
        if "+" in src:
            for part in [p.strip() for p in src.split("+")]:
                if (part and not part.startswith("-")
                        and part not in roots and part not in NON_ROOT_PARTS):
                    found.append(("dangling part", e["root"],
                                  f"Built from '{part}', which is not a root anywhere in the lexicon."))
        if not src:
            found.append(("no derivation", e["root"], "No Derivation cell, so the root cannot be checked."))
        if not str(e["meaning"] or "").strip():
            found.append(("no meaning", e["root"], "No Meaning cell."))
    for root, note in OVERRIDE.items():
        found.append(("lexical override", root, note))
    for root, vowel in PAD_VOWEL.items():
        found.append(("lexical pad vowel", root,
                      f"Takes -{vowel} rather than the default -a. No rule predicts this."))

    r = write_lines(s, [
        ("", False),
        ("OPEN ISSUES. GENERATED by build.py on every run.", True),
        ("", False),
        ("The top table is DETECTED from the data on every build. If a row appears here, the build found it, not a person.", False),
        ("The bottom table is decisions recorded as open. Those are maintained by hand.", False),
        ("", False),
        (f"DETECTED: {len(found)} items", True),
    ])
    for j, h in enumerate(["Kind", "Root", "Detail"], start=1):
        s.cell(row=r, column=j, value=h).font = HEAD
    r += 1
    for kind, root, detail in sorted(found):
        for j, v in enumerate((kind, root, detail), start=1):
            c = s.cell(row=r, column=j, value=v)
            c.font = BODY
            c.alignment = Alignment(wrap_text=(j == 3), vertical="top")
        if kind in ("data error", "dangling part"):
            s.cell(row=r, column=1).fill = RED
        else:
            s.cell(row=r, column=1).fill = GREY
        r += 1
    r += 2
    s.cell(row=r, column=1, value=f"OPEN DECISIONS: {len(OPEN_DECISIONS)}").font = HEAD
    r += 1
    for j, h in enumerate(["Item", "Status", "Detail"], start=1):
        s.cell(row=r, column=j, value=h).font = HEAD
    r += 1
    for item, status, detail in OPEN_DECISIONS:
        for j, v in enumerate((item, status, detail), start=1):
            c = s.cell(row=r, column=j, value=v)
            c.font = BODY
            c.alignment = Alignment(wrap_text=(j == 3), vertical="top")
        r += 1
    return found


def build_grammar(wb):
    """Add the possessive, and fix the ORANGE rows, which name the profession
    root as oran. That is now the COLOUR. The profession is orang."""
    g = wb["Grammar"]
    fixed = 0
    for row in g.iter_rows():
        if row[0].value == "ORANGE":
            for c in row:
                if isinstance(c.value, str) and c.value.strip() == "oran":
                    c.value = "orang"
                    c.font = BODY
                    fixed += 1
    # De-duplicate first. Earlier builds appended without checking, so the
    # numeral and substrate rows landed twice.
    seen, drop = set(), []
    for r in g.iter_rows():
        key = (r[0].value, r[1].value, r[2].value)
        if key in seen and r[0].value:
            drop.append(r[0].row)
        seen.add(key)
    for n in reversed(drop):
        g.delete_rows(n)
    # ploh is nine, and nine is ONE BLOCK. The Block root moved to pluh, so
    # the numeral has to follow it or the two stop agreeing.
    for r in g.iter_rows():
        if r[0].value == "NUMERALS" and isinstance(r[2].value, str) and "ploh" in r[2].value:
            r[2].value = r[2].value.replace("ploh", "pluh")
            r[1].value = "pluh" if r[1].value == "ploh" else r[1].value
            r[2].font = BODY
    have = {(r[0].value, r[1].value) for r in g.iter_rows()}
    rows = [
        ("POSSESSIVE", "-na",
         "Marks the POSSESSOR, not the possessed. vilbonka is a village bank, a kind of bank. vilana bonka is the "
         "village's bank, that specific one. The suffix is what makes the phrase not a compound."),
        ("POSSESSIVE", "Why -na",
         "Consonant-initial, because 74 of 215 roots end in a vowel and every other suffix is vowel-initial. A "
         "vowel-initial possessive would hit hiatus on a third of the lexicon and the junction rule would eat the "
         "root's own ending."),
        ("POSSESSIVE", "Affix phonology",
         "Affixes do NOT trigger the compound junction rule. If they did, bid plus -na would delete the final stop and "
         "give bina, destroying the root. Compounding and affixation have different phonology."),
        ("POSSESSIVE", "Scope",
         "Ownership only. No case, no agreement, no other relation. It is the one relation villagers demonstrably "
         "understand, which is why it is the only grammar there is."),
    ]
    for row in rows:
        if (row[0], row[1]) in have:
            continue
        g.append(row)
        for c in g[g.max_row]:
            c.font = BODY
            c.alignment = Alignment(wrap_text=True, vertical="top")
    c1_key = ("COMPOUNDING", "C1 place names")
    if c1_key not in {(r[0].value, r[1].value) for r in g.iter_rows()}:
        g.append((
            "COMPOUNDING", "C1 place names",
            "Place names use C1 seam repairs from root_converter.junction(), but the modifier clip is OFF. Both source roots remain visible; reroll a pair if the repaired output is phonotactically illegal."
        ))
        for cell in g[g.max_row]:
            cell.font = BODY
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    return fixed


def pin_drift_order(wb):
    """Keep the interaction order explicit in the maintained Substrate sheet."""
    s = wb["Substrate"]
    for row in s.iter_rows():
        if row[0].value == "DRIFT" and row[1].value == "Order":
            row[2].value = (
                "Biome rule first, then drift. Within drift, apply the selected D-rules "
                "in ascending rule index; D3 feeding D9 is not D9 feeding D3."
            )
            row[2].font = BODY
            row[2].alignment = Alignment(wrap_text=True, vertical="top")
            return
    s.append((
        "DRIFT", "Order",
        "Biome rule first, then drift. Within drift, apply the selected D-rules in ascending rule index."
    ))
    for cell in s[s.max_row]:
        cell.font = BODY
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def parse_args():
    import argparse
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--in", dest="src", default=SRC, help="workbook to read")
    p.add_argument("--out", dest="out", default=OUT,
                   help="workbook to write, or the word 'same' to write back over the input")
    a = p.parse_args()
    return a.src, (a.src if a.out == "same" else a.out)


def main():
    src, out = parse_args()
    verified_compounds = verify_compounds()
    wb = load_workbook(src)
    entries, seen = read_master(wb)
    before = len(entries)
    added = add_incoming(entries, seen, wb["Concept Roots"])

    ws = wb["Concept Roots"]
    hdr = None
    for row in ws.iter_rows():
        if row[0].value == "Root":
            hdr = row
            break
    if hdr:
        for off, name in ((10, "Section"), (11, "Village naming")):
            ws.cell(row=hdr[0].row, column=off + 1, value=name).font = HEAD
    by = {e["root"]: e for e in entries}
    for row in ws.iter_rows():
        v = [c.value for c in row]
        if v[0] in (None, "Root") or not v[3]:
            continue
        e = by.get(str(v[0]).strip())
        if not e:
            continue
        ws.cell(row=row[0].row, column=11, value=e["section"]).font = BODY
        ws.cell(row=row[0].row, column=12,
                value="no" if e["root"] in NO_VILLAGE_NAME else "yes").font = BODY

    refreshed = refresh_compound_parts(wb, entries)
    dstats, dblocked, dhom = build_dialects(entries, wb)
    nmissing, ngone = build_names(wb, entries)
    sfixed = build_strata(wb)
    dfixed = build_dialect_examples(wb, entries)
    build_index(wb, entries)
    build_sound_system(wb)
    build_phonotactics(wb, entries)
    counts, s, r = build_regeneration(wb, entries)
    found = build_adopt(wb, entries, s, r + 3)
    fixed = build_grammar(wb)
    pin_drift_order(wb)

    # The banner is idempotent. Inserting unconditionally added a duplicate row
    # on every run, and the sheet grew a banner per build.
    if "Design Notes" in wb.sheetnames:
        ws = wb["Design Notes"]
        banner = ("RATIONALE AND HISTORY ONLY. No root is defined here. "
                  "Concept Roots is the single source of truth.")
        # Earlier builds inserted unconditionally, so the sheet accumulated one
        # banner per run. Collapse them all, then write exactly one.
        stale = [r for r in range(ws.max_row, 0, -1)
                 if str(ws.cell(row=r, column=1).value or "").strip() == banner]
        for r in stale:
            ws.delete_rows(r)
        ws.insert_rows(1)
        ws.cell(row=1, column=1, value=banner).font = HEAD
        if stale:
            print(f"  Design Notes banners collapsed  {len(stale)} -> 1")

    # Status is deleted. It was a change log of superseded proposals: mend >
    # menth, bane > pame, "I moved Block to ploh", none of which happened or
    # happened differently, and it contradicted almost every other sheet. The
    # Build Report carries the live state and Design Notes carries the history.
    if "Status" in wb.sheetnames:
        del wb["Status"]

    wb.save(out)
    print(f"  roots before consolidation  {before}")
    print(f"  moved in from other sheets  {added}")
    print(f"  total                       {len(entries)}")
    for k, v in counts.items():
        print(f"  {k:<20} {v}")
    print(f"  compound parts fixed  {len(refreshed)}")
    for r in refreshed:
        print(f"    {r[0]:<10} {r[1]:<18} -> {r[2]}")
    print("  dialect forms generated:")
    for tag in dstats:
        print(f"    {tag}  {dstats[tag]:>3} forms, {dblocked[tag]:>3} blocked by the constraints")
    print(f"  blocked as homonyms   {len(dhom)}  {dhom}")
    print(f"  professions missing a root  {nmissing}")
    print(f"  redundancy roots missing    {ngone}")
    print(f"  strata cells rebuilt  {sfixed}")
    print(f"  dialect examples      {dfixed}")
    print(f"  detected issues       {len(found)}")
    print(f"  grammar cells fixed   {fixed}")
    print(f"  C1 compounds verified  {verified_compounds}")
    print(f"  written to {out}")


if __name__ == "__main__":
    main()
