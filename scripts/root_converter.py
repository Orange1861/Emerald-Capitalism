#!/usr/bin/env python3
"""
Emerald Capitalism: villager root converter (revision 2).

Turns an English source word into a villager root by applying the sound
rules in order, then checks the output against the phonotactic constraints.

Usage:
    python3 root_converter.py golem 2
    python3 root_converter.py "grindstone" 3
    python3 root_converter.py snow 3 --metathesis
    python3 root_converter.py --audit                (phonotactics of every recorded root)
    python3 root_converter.py --regenerate           (derive every root, report cost)
    python3 root_converter.py --solve anva Anvil 3   (cheapest derivation of one root)

Stratum 2a = deep integration, pre-shift   -> S0-S7 plus the S8 shift
Stratum 2b = deep integration, post-shift  -> S0-S7, no shift, S10a still holds
Stratum 3  = late integration              -> S0-S7, no shift
Stratum 4  = collapse-era                  -> S0, S1 only
Stratum 5  = post-collapse                 -> no sound change at all

THREE RULES ARE SPORADIC and never fire automatically:
    --aphaeresis  S5, delete a word-initial consonant
    --paragoge    S6, add the final -a
    --metathesis  S9, transpose a liquid or glide with an adjacent vowel
    --breaking    S7, i > ai before a resonant

WHAT CHANGED IN REVISION 3
--------------------------
S0b VOWEL LENGTH. English tense vowels now map to a two-segment nucleus:
ei, ii, ai, oo, uu, au. A five-vowel inventory collapsed Boat with Bottle,
Banner with Bane and Golem with Cauldron, and the collisions were an
artefact of the converter rather than of the sources. The lexicon voted the
other way (11 of 14 /oU/ roots used plain o), so this is a design decision
that overrides the recorded data, not a fit to it. Long nuclei are SINGLE
segments, like ch, so Iron gives air and not a-ir.

S3b NARROWED. It no longer erases a LEGAL coda cluster. Sharpness and Shard
both came out shar, which discarded a contrast the sources made. What
remains of S3b is th-stopping.

S3c COMPENSATORY RAISING, automatic. Monosyllabic wear moved off the coda
and onto the vowel, where it changes the word without destroying a contrast.
Chest gives chist, Skulk skolk, Swamp swomp. A long nucleus does not raise.

wl IS LEGAL, as an onset and as a coda. Metathesis produces it (Wool gives
wlo) and the lexicon carries it, so the constraint table has to allow it.

z IS LEGAL AT STRATA 4 AND 5 ONLY, and survives G2P there rather than
mapping to s. Those strata already sit outside the cluster rules for the
same reason: they froze almost unworn. Blaze gives bleiz at stratum 4 and
blais anywhere else.

P10 MINIMUM CONTENT. No reduction may leave fewer than three segments, and
S3 now trims a cluster from the right one segment at a time instead of
collapsing to its first member. Emerald and Empty stopped both being em.

EFFECT: re-deriving every borrowed root from its recorded source, stratum
and sporadic rules, collisions fall from 8 to 3, and all three remaining
ones are two senses of a single English source (Sculk, Smith, Emerald),
which is a real homophone rather than a converter failure.

WHAT CHANGED IN REVISION 2
--------------------------
1. Everything runs on SEGMENTS, not letters. tokenize() reads ch, sh, th,
   ng and kw as single sounds. Length tests, cluster tests, metathesis and
   the minimal-word floor now count sounds.
2. Source preprocessing splits into three named passes: source_pronunciation
   lookup, english_g2p, then map_to_inventory. A strange output can now be
   traced to English spelling or to a villager sound change, not both.
3. Stress lives in one place. s0_stress() consults the override table and
   the Latinate heuristic and returns both the syllable index and the
   truncation mode. convert() no longer decides stress on its own.
4. Strata are first-class. The STRATA table states what each one does.
   Pipelines are declared as ordered lists of Rule objects, so the relative
   order of S5, S8, S7, S6, P9, S3 is data rather than control flow.
5. Rules carry a category: historical, sporadic, repair, or synchronic.
6. The solver ranks derivations by COST and returns the cheapest one, so a
   root that needs three sporadic rules plus a stratum change is visibly
   more expensive than one that falls out regularly.
7. Paragoge adds -a. All 38 roots recorded as taking S6 end in -a, so the
   harmony guess was overwriting attested data. harmony_vowel() survives as
   a generator suggestion only.
8. S3d raising is sporadic and OFF. Firing it automatically on every
   one-syllable source cost 23 exact reproductions and bought 2.
9. The minimal-word floor is three SEGMENTS, not four letters.
10. legal() counts syllables with nuclei(), checks the whole segment
   inventory rather than four banned letters, and knows the stratum.
11. Retired rules S2 and S4 moved to retired_rules.py. Dead constants gone.

MEASURED EFFECT on exact reproduction of the recorded lexicon, deriving each
root from its recorded source, stratum and sporadic rules:

    revision 1                                   23 of 169
    + paragoge adds -a instead of harmonising    38
    + S3d raising sporadic rather than automatic 61
    + soft c, source pronunciation table, floor  63

The remaining gap is not a converter problem. --regenerate reports 88 roots
that no combination of rules produces at any price. Those were coined by
hand before the sound system existed, and they are the list to work through.
"""

import re
import sys

VOWELS = set("aeiou")
RESONANTS = set("lrmn")
LIQUID_GLIDE = set("lrwy")

OBSTRUENTS = set("ptkbdgfvsh")
LIQUIDS_GLIDES = set("lrwy")
NASALS = set("mn")

# The full segment inventory. j is permitted. c, q, x, y and z are not
# segments at all: G2P maps every one of them to something the language has,
# so a root containing one is a spelling that the converter could never
# produce. Bare c is the common case, and revision 1 missed it because it
# only checked x, y, z and q.
CONSONANT_SEGMENTS = set("bdfghjklmnprstvw")

# z SURVIVES IN THE COLLAPSE-ERA VOCABULARY ONLY. Strata 4 and 5 froze
# almost unworn, so they keep source material the settled phonology never
# took in. Everywhere else z still maps to s in G2P. This is the same
# exemption strata 4 and 5 already have from the cluster rules.
LATE_ONLY_SEGMENTS = set("z")


# ============================================================ SEGMENTS
# One letter is not one sound. ch, sh, th, ng and kw are each a single
# segment, and every rule that counts, reduces, reorders or measures has
# to count them once. Revision 1 knew this in three places (S3b exempted
# them, S8 skipped them, legal_onset special-cased them) and forgot it
# everywhere else, which is why the minimal-word floor padded chest but
# not shard.

DIGRAPHS = ("ch", "sh", "th", "ng", "kw",
            "ei", "ih", "ii", "ai", "oo", "uu", "au", "oi")
DIGRAPH_ONSETS = ("sh", "th", "ch", "kw")

# A long nucleus is ONE segment, exactly as ch is. Without this, syllabify()
# splits air into a-ir and S1 truncates Iron to ai instead of air.
LONG_NUCLEI = ("ei", "ih", "ii", "ai", "oo", "uu", "au", "oi")


def tokenize(word):
    """Split a string into phonological segments. chest -> ch e s t."""
    segs, i = [], 0
    while i < len(word):
        if word[i:i + 2] in DIGRAPHS:
            segs.append(word[i:i + 2])
            i += 2
        else:
            segs.append(word[i])
            i += 1
    return segs


def untokenize(segs):
    return "".join(segs)


def is_vowel(seg):
    return seg in VOWELS or seg in LONG_NUCLEI


def seg_len(word):
    """How many sounds, not how many letters."""
    return len(tokenize(word))


def onset_of(word):
    """Leading consonant segments."""
    segs = tokenize(word)
    out = []
    for s in segs:
        if is_vowel(s):
            break
        out.append(s)
    return out


def coda_of(word):
    """Trailing consonant segments."""
    segs = tokenize(word)
    out = []
    for s in reversed(segs):
        if is_vowel(s):
            break
        out.insert(0, s)
    return out


# ==================================================== SOURCE PREPROCESSING
# Three passes, kept apart on purpose.
#   source_pronunciation()  a lookup for words English spells dishonestly
#   english_g2p()           the heuristic fallback
#   map_to_inventory()      villager sound inventory, not English at all
# Revision 1 ran all three inside normalise(), so a wrong output could not
# be blamed on the right stage.

# Words the heuristic gets wrong. The intended vocabulary is finite, so a
# lookup is more reliable than another spelling rule. Values are already
# in one-letter-per-sound form and skip english_g2p entirely.
#
# KEEP THIS TABLE TO PRONUNCIATION. A compound whose medial e is doing
# syllabic work belongs to morphology, not spelling: an entry of
# lodestone > lodston deleted the syllable break and produced lodstna,
# where the heuristic alone gives lod > loda, the attested root.
SOURCE_PRONUNCIATION = {
    "eye": "ai",
    "fence": "fens",
    "price": "pris",
    "furnace": "furnas",
    "guard": "gard",
    "sign": "sin",
    "ruin": "ruin",
    "ruined": "ruind",
    "runes": "run",
    "drowned": "dround",
    "acacia": "akasha",
    "player": "plaer",
    "geode": "geod",
}

CONSONANT_DIGRAPHS = [("tch", "ch"), ("dge", "j"), ("ck", "k"), ("ph", "f"),
                      ("wh", "w"), ("qu", "kw"), ("gh", ""), ("kn", "n")]
VOWEL_DIGRAPHS = [("eau", "o"), ("ough", "o"), ("augh", "o"), ("igh", "i"),
                  ("oa", "o"), ("oe", "o"), ("oo", "u"), ("ou", "o"),
                  ("ai", "a"), ("ay", "a"), ("au", "o"), ("aw", "o"),
                  ("ea", "e"), ("ee", "e"), ("ei", "e"), ("ey", "e"),
                  ("ie", "i"), ("ue", "u"), ("ui", "u"), ("eu", "u"), ("ew", "u")]

# Compound sources whose FIRST element is generic and carries no meaning
# villagers cared about. English stresses these initially (MINEcart,
# MINEshaft), so S0 lands on the wrong element and the root comes out of
# the modifier instead of the head. Strip them before stress is assigned.
LOW_STRESS_PREFIXES = ("mine", "mini")


def english_g2p(word):
    """English spelling to one letter per sound. English side only."""
    w = word.lower().strip()
    for a, b in CONSONANT_DIGRAPHS:
        w = w.replace(a, b)
    if len(w) > 3 and w.endswith("e") and w[-2] not in VOWELS:
        w = w[:-1]                       # silent final e: bale > bal
    for a, b in VOWEL_DIGRAPHS:
        w = w.replace(a, b)
    # Word-final ow is /oU/ and the w is not pronounced. S0b now carries that
    # vowel as a long nucleus, so the w has no work left to do. This does
    # cost snwo, which was only derivable when the silent w survived to be
    # metathesised.
    if w.endswith("ow"):
        w = w[:-2] + "o"
    # SOFT C. English c is /s/ before e, i and y and /k/ elsewhere. Revision
    # 1 mapped every c to k, which turned fence into fenk and price into
    # prik. Both then failed to derive their recorded roots.
    w = re.sub(r"c(?=[eiy])", "s", w)
    w = re.sub(r"c(?!h)", "k", w)
    w = re.sub(r"([^aeiou])\1", r"\1", w)     # doubled consonant = one segment
    return w


def map_to_inventory(word, stratum=None):
    """
    Villager inventory, not English. Letters the language does not carry
    are converted, not merely flagged. j is KEPT, and z is kept at strata 4
    and 5, where the collapse-era vocabulary froze before the phonology
    finished absorbing it. Blaze gives bleiz there and blais elsewhere.
    """
    w = word
    pairs = [("x", "ks"), ("q", "k"), ("y", "i")]
    if str(stratum) not in ("4", "5"):
        pairs.append(("z", "s"))
    for a, b in pairs:
        w = w.replace(a, b)
    return w


def source_morphology(word):
    """Discard a generic leading element so S0 falls on the head."""
    for p in LOW_STRESS_PREFIXES:
        if word.startswith(p) and len(word) - len(p) >= 3:
            return word[len(p):]
    return word


def normalise(word, stratum=None):
    """Full source preprocessing. Kept as the single public entry point."""
    raw = word.lower().strip()
    if raw in SOURCE_PRONUNCIATION:
        w = SOURCE_PRONUNCIATION[raw]
    else:
        w = english_g2p(raw)
    w = map_to_inventory(w, stratum)
    w = apply_s0b(w, raw, stratum)
    w = source_morphology(w)
    return w


# ==================================================== VOWEL LENGTH (S0b)
# The language distinguishes a long nucleus from a short one. English does
# too, and spelling hides it: Boat and Bottle differ only in the vowel, and a
# five-vowel inventory collapsed them. Three root collisions traced to this
# one gap (Boat/Bottle, Banner/Bane, Golem/Cauldron), so the contrast is
# carried into the villager form as a two-segment nucleus.
#
#   /eI/ bale  -> ei     /i:/ green -> ii     /aI/ night -> ai
#   /oU/ boat  -> oo     /u:/ bloom -> uu     /aU/ count -> au
#
# The table is keyed on the SOURCE word because English spelling cannot be
# trusted to tell you which vowel it is. It was generated from the CMU
# Pronouncing Dictionary and then hand-checked for the compounds the
# dictionary does not carry. To add a word, look up its stressed vowel and
# add a line; if cmudict is installed, unknown words fall through to it.
TENSE_NUCLEUS = {
    "acacia": "ei", "bale": "ei", "bamboo": "uu", "bane": "ei",
    "binding": "ai", "blaze": "ei", "bloom": "uu", "boat": "oo",
    "bone": "oo", "bonemeal": "oo", "cave": "ei", "count": "au",
    "daylight": "ei", "deal": "ii", "diamond": "ai", "drowned": "au",
    "dune": "uu", "eye": "ai", "frame": "ei", "gate": "ei",
    "geode": "ii", "glowstone": "oo", "golem": "oo", "green": "ii",
    "grindstone": "ai", "hoe": "oo", "iron": "ai", "loam": "oo",
    "lodestone": "oo", "loom": "uu", "mason": "ei", "night": "ai",
    "node": "oo", "oak": "oo", "omen": "oo", "outpost": "au",
    "player": "ei", "price": "ai", "rail": "ei", "rain": "ei",
    "reach": "ii", "reeds": "ii", "roots": "uu", "ruin": "uu",
    "ruined": "uu", "runes": "uu", "scribe": "ai", "sign": "ai",
    "smoker": "oo", "snow": "oo", "solar": "oo", "spruce": "uu",
    "stone": "oo", "survive": "ai", "terrain": "ei", "totem": "oo",
    "trade": "ei", "trial": "ai", "trident": "ai", "wheat": "ii",
}


_CMUDICT = None


def _cmudict():
    """Load the pronouncing dictionary ONCE. Calling cmudict.dict() per word
    rebuilds a 130,000 entry dictionary every time, which turned a 300 word
    test into a job that never finished."""
    global _CMUDICT
    if _CMUDICT is None:
        try:
            import cmudict
            _CMUDICT = cmudict.dict()
        except ImportError:
            _CMUDICT = {}
    return _CMUDICT


def _tense_nucleus(word):
    """Which long nucleus this source takes, or None if its vowel is short."""
    w = word.lower().strip()
    if w in TENSE_NUCLEUS:
        return TENSE_NUCLEUS[w]
    pr = _cmudict().get(w)
    if not pr:
        return None
    stressed = next((p[:-1] for p in pr[0] if p[-1:] == "1"), None)
    return {"EY": "ei", "IY": "ii", "AY": "ai", "OW": "oo",
            "UW": "uu", "AW": "au", "OY": "oi"}.get(stressed)


def long_i_mark(word, stratum=None):
    """
    Long /i:/ is written ih, EXCEPT where an h is already in play, and then
    it is ii.

    h is a live consonant here: twelve roots carry a standalone one, and the
    shift turns every k into h, so pluh and hah are u plus h and a plus h. In
    a word that already has an h, a reader has no way to tell which h is a
    consonant and which is a length mark, so the doubled form is used
    instead. The shift counts too, because a k at stratum 2a BECOMES an h
    later in the derivation: Keep would give hihp, so it gives hiip.
    """
    segs = tokenize(word)
    if any(s == "h" for s in segs):
        return "ii"
    if stratum and STRATA.get(str(stratum), {}).get("shift") and any(s == "k" for s in segs):
        return "ii"
    return "ih"


def apply_s0b(word, source, stratum=None):
    """
    S0b. Lengthen the stressed nucleus when the English source has a tense
    vowel. Runs on the G2P output, before stress and truncation, because
    every later rule counts segments and a long nucleus is two of them.
    """
    nucleus = _tense_nucleus(source)
    if nucleus == "ii":
        nucleus = long_i_mark(word, stratum)
    if not nucleus:
        return word
    segs = tokenize(word)
    for i, s in enumerate(segs):
        if is_vowel(s):
            if i + 1 < len(segs) and is_vowel(segs[i + 1]):
                return word          # already a long nucleus, leave it
            segs[i] = nucleus
            return untokenize(segs)
    return word

# ==================================================== SYLLABLES AND STRESS

def syllabify(word):
    """
    Split into syllables. Two adjacent DIFFERENT vowels are two nuclei
    (di-a-mond); a doubled vowel is one (bo-ok > book).

    This is deliberately CODA-HEAVY: golem parses as gol-em, not go-lem.
    S1 truncation is measured against that parse, so do not replace it with
    a maximal-onset syllabifier without re-measuring S1.
    """
    segs = tokenize(word)
    sylls, cur, seen = [], "", False
    for i, seg in enumerate(segs):
        nxt = segs[i + 1] if i + 1 < len(segs) else ""
        cur += seg
        if is_vowel(seg):
            if seen and is_vowel(nxt) and nxt != seg:
                sylls.append(cur); cur, seen = "", False; continue
            seen = True
            if is_vowel(nxt) and nxt != seg:
                sylls.append(cur); cur, seen = "", False
        elif seen and is_vowel(nxt):
            sylls.append(cur); cur, seen = "", False
    if cur:
        if sylls and not seen:
            sylls[-1] += cur
        else:
            sylls.append(cur)
    return sylls or [word]


def nuclei(word):
    """
    S10b DIPHTHONG INTEGRITY. Count vowel NUCLEI, where a vowel sequence is
    ONE nucleus. syllabify() splits different-vowel sequences because S1b
    needs that split to produce dia; nuclei() is the speaker's count, and it
    is the one the stratum-2 shape rule is measured against.
    """
    segs = tokenize(word)
    n, i = 0, 0
    while i < len(segs):
        if is_vowel(segs[i]):
            n += 1
            while i + 1 < len(segs) and is_vowel(segs[i + 1]):
                i += 1
        i += 1
    return n


# S0 override table. Keyed on the NORMALISED form, so add entries by running
# normalise() first. Revision 1 held eleven entries whose value was 0, which
# is what s0_stress already returns, and five of those keys could never match
# a normalised form. The mechanism is kept; the inert entries are gone.
STRESS_OVERRIDE = {}

LATINATE_PRE_STRESS = ("tion", "sion", "ity", "ic", "ical", "ial", "ian",
                       "ious", "eous", "ify", "ual", "uous")


def latinate_stress(word):
    """
    S0, Latinate branch. English stresses the syllable BEFORE -tion, -ity,
    -ic, -ial and friends, and the antepenult in most other Latinate
    polysyllables. Defaulting to the first syllable gave inFLAtion > in and
    colLATeral > col, which loses the part of the word that carries meaning.
    Returns a syllable index, or None if the word is not Latinate-shaped.
    """
    s = syllabify(word)
    if len(s) < 3:
        return None
    for suf in LATINATE_PRE_STRESS:
        if word.endswith(suf):
            n = len(syllabify(suf))
            idx = len(s) - n - 1
            return idx if idx >= 0 else 0
    if seg_len(word) >= 7:
        return len(s) - 3        # antepenult
    return None


def s0_stress(word):
    """
    S0. English primary stress, decided in ONE place.
    Returns (syllable_index, mode).

    mode "head"     -> S1 truncates to the stressed syllable.
    mode "through"  -> S1 keeps everything up to and through it. Latinate
                       words need this: coda-heavy syllabification does not
                       line up with English stress there, and truncating
                       from the stress gave inflation > atha.
    """
    if word in STRESS_OVERRIDE:
        return STRESS_OVERRIDE[word], "head"
    lat = latinate_stress(word)
    if lat is not None and lat > 0:
        return lat, "through"
    return 0, "head"


# ============================================================ SOUND RULES
# Every rule takes a word and returns a word. Ordering lives in the stratum
# pipelines below, not in these functions.

def apply_s1(word, stress, mode="head"):
    """
    S1. Truncation. Syllabification is coda-heavy, so the stressed syllable
    already carries the consonant that follows its vowel (go-lem parses as
    gol-em). Three cases, decided by the syllable after it:

    a) DIFFERENT vowel, closed head -> the head is already complete. Stop.
         Golem (gol-em) > gol   Anvil (anv-il) > anv   Warden (ward-en) > ward
    b) DIFFERENT vowel, OPEN head -> true hiatus (S1b). Absorb it.
         Diamond (di-am-ond) > dia
    c) SAME vowel -> keep both syllables and let bounded syncope shorten them.
         Emerald (em-er-ald) > emer > emr
    """
    s = syllabify(word)
    if mode == "through":
        return "".join(s[:stress + 1])
    if stress >= len(s):
        return word
    head = s[stress]
    if stress + 1 >= len(s):
        return head
    nxt = s[stress + 1]
    onset = "".join(onset_of(nxt))
    if onset:
        return head + onset
    last_vowel = next((c for c in reversed(tokenize(head)) if is_vowel(c)), "")
    if nxt[0] == last_vowel:
        # Case c, with bounded syncope folded in. Drop the next syllable's
        # vowel and keep its consonants, but only when that leaves a SINGLE
        # consonant segment. This is retired S2 scoped to the one place it
        # earned its keep.
        #   Emerald: em + r        -> emr  (one segment, kept)
        #   Binding: bind + ng     -> bind (two segments, fall back to head)
        rest = [c for c in tokenize(nxt) if not is_vowel(c)]
        if len(rest) == 1:
            return head + rest[0]
        return head
    if not is_vowel(tokenize(head)[-1]):
        return head                           # case a
    return head + nxt[0]                      # case b


def apply_s3(word):
    """
    S3. Cluster reduction, driven by the constraint table.
    A final consonant cluster reduces to its first segment ONLY if the full
    cluster is illegal. Crafting > craft > craf (ft illegal). Mending >
    mend, kept (nd legal). Runs after S6, so a cluster that gained the -a
    is no longer a coda and survives: anv > anva.
    """
    coda = coda_of(word)
    if len(coda) < 2 or untokenize(coda) in LEGAL_CODA_2:
        return word
    stem = word[:len(word) - len(untokenize(coda))]
    # Trim the coda from the RIGHT one segment at a time and stop at the
    # first legal cluster, rather than collapsing straight to the first
    # segment. emfth gives emf, not em, so Empty keeps its shift.
    # P10 MINIMUM CONTENT: a reduction may not leave fewer than three
    # segments. Emerald and Empty both truncated to em, which is a collision
    # the rules created rather than one the sources had.
    for n in range(len(coda) - 1, 0, -1):
        trial = stem + untokenize(coda[:n])
        if seg_len(trial) < MIN_CONTENT:
            break
        if untokenize(coda[:n]) in LEGAL_CODA_2 or n == 1:
            return trial
    return word


def apply_s3b(word):
    """
    S3b. Monosyllabic coda wear, NARROWED.
    It used to reduce any final cluster to its first segment, legality be
    damned. That erased contrasts the source made: Sharpness and Shard both
    came out shar, because rp and rd are both legal and both got cut. A
    collision between two roots that each went through the rules means the
    converter is discarding a distinction, so the rule now leaves a LEGAL
    cluster alone and S3c carries the wear instead.

    What survives here is th-stopping, which is a segment change rather than
    a cluster reduction: Path > pat, Smith > smit.
    """
    coda = coda_of(word)
    keep = len(word) - len(untokenize(coda))
    if coda == ["th"]:
        return word[:keep] + "t"      # th-stopping: Path > pat, Smith > smit
    if len(coda) < 2:
        return word                   # a digraph is one segment, not a cluster
    if untokenize(coda) in LEGAL_CODA_2:
        return word                   # a legal cluster is contrast, not wear
    return word[:keep] + coda[0]


RAISE = {"a": "o", "e": "i", "o": "u", "u": "o", "i": "e"}


def apply_s3c(word):
    """
    S3c. Compensatory raising. AUTOMATIC for a one-syllable source.
    The stem vowel shifts one step: a > o, e > i, o > u, u > o, i > e.

    This is now the whole of monosyllabic wear. A one-syllable source loses
    nothing to truncation, so something has to mark it as borrowed rather
    than copied, and S3b no longer does that job without destroying coda
    contrast. Raising changes the word while keeping every consonant.

    Chest > chist. Skulk > skolk. Swamp > swomp. Lush > losh.

    A LONG nucleus does not raise. It is already heavy enough to count as
    the distinctive material, and shortening it would undo the contrast S0b
    exists to carry.
    """
    segs = tokenize(word)
    for i, seg in enumerate(segs):
        if is_vowel(seg):
            if seg in LONG_NUCLEI:
                return word
            segs[i] = RAISE.get(seg, seg)
            return untokenize(segs)
    return word


def apply_s5(word):
    """
    S5. Aphaeresis. Delete a single word-initial consonant.
    SPORADIC. Only three attestations (husk, punch, silk), so this never
    fires automatically. Pass --aphaeresis to apply it.
    """
    segs = tokenize(word)
    if len(segs) > 2 and not is_vowel(segs[0]) and is_vowel(segs[1]):
        return untokenize(segs[1:])
    return word


def degeminate(word):
    """A doubled segment simplifies. Bell > bel, so S6 gives bela not bella."""
    segs = tokenize(word)
    out = [segs[0]]
    for seg in segs[1:]:
        if seg != out[-1]:
            out.append(seg)
    return untokenize(out)


def harmony_vowel(word):
    """
    NOT USED BY THE CONVERTER. Kept as a suggestion for the generator.

    The hypothesis was that the paragogic vowel agrees with the stem vowel:
    back stem takes -o, high front stem takes -i, everything else -a. The
    lexicon rejects it. All 38 roots recorded as taking S6 end in -a, and
    harmony predicts only 14 of them. The i, o and u finals elsewhere in the
    lexicon come from the source vowel, not from paragoge. If a root ever
    takes a non-a paragogic vowel, record it on that root as lexical data
    rather than deriving it.
    """
    v = next((c for c in reversed(word) if c in VOWELS), "a")
    if v in "ou":
        return "o"
    if v == "i":
        return "i"
    return "a"


def apply_s6(word, vowel=None):
    """
    S6. Paragogic -a. SPORADIC at every stratum.
    Measured across all 176 roots: 42 percent of stratum 2, 48 percent of
    stratum 3 and 23 percent of stratum 4 end in -a. There is no stratum
    rule and no phonological conditioning. Apply it by hand.
    """
    # The syllable gate matches the phonotactic ceiling rather than a
    # separate hard-coded 2. In practice S1 has already cut everything below
    # the ceiling by the time S6 runs, so this gate never fires today; it is
    # here so that loosening the ceiling loosens both together.
    if word and not is_vowel(tokenize(word)[-1]) and nuclei(word) <= MAX_NUCLEI:
        return degeminate(word) + (vowel or "a")
    return word


def apply_s7(word):
    """
    S7. Breaking. i > ai before a resonant (l, r, m, n). SPORADIC.
    Iron > ir > (S7) air > (S6) aira.
    """
    segs = tokenize(word)
    out = []
    for i, seg in enumerate(segs):
        nxt = segs[i + 1] if i + 1 < len(segs) else ""
        prev = segs[i - 1] if i > 0 else ""
        if seg == "i" and nxt in RESONANTS and (i == 0 or not is_vowel(prev)):
            out.append("ai")
        else:
            out.append(seg)
    return untokenize(out)


S8 = {"p": "f", "t": "th", "k": "h", "b": "p", "d": "t", "g": "k", "v": "f"}


def apply_s8(word):
    """S8. The shift. Stratum 2a only. Digraphs are single segments and are
    already at their shifted value, so they pass through."""
    return untokenize([S8.get(s, s) for s in tokenize(word)])


def s9_sites(word):
    """Every position where a liquid/glide and an adjacent vowel could swap."""
    segs = tokenize(word)
    out = []
    for i in range(len(segs) - 1):
        a, b = segs[i], segs[i + 1]
        if (a in LIQUID_GLIDE and is_vowel(b)) or (is_vowel(a) and b in LIQUID_GLIDE):
            out.append(i)
    return out


def apply_s9(word, site=0):
    """
    S9. Metathesis. SPORADIC.
    A word can have SEVERAL eligible sites, and which one metathesises is a
    property of the word, not of the rule. `site` indexes s9_sites(); the
    solver tries them all. Wool > wul has two sites: swapping at 0 gives
    uwl, at 1 gives wlu. Only trying the leftmost missed wlu entirely.
    """
    sites = s9_sites(word)
    if not sites or site >= len(sites):
        return word
    i = sites[site]
    segs = tokenize(word)
    segs[i], segs[i + 1] = segs[i + 1], segs[i]
    return untokenize(segs)


MIN_CONTENT = 3
MAX_NUCLEI = 3


def apply_p10(word):
    """
    P10. MINIMUM CONTENT repair. A cluster that S3 was forbidden to reduce is
    still illegal, so it takes the paragogic vowel and stops being a coda at
    all. Emerald: emr, which S3 may not cut to em, becomes emra.
    """
    coda = coda_of(word)
    if len(coda) >= 2 and untokenize(coda) not in LEGAL_CODA_2:
        return word + "a"
    return word


def apply_s11(word):
    """
    S11. Final unstressed -er becomes -a. STRATUM 4 ONLY.

    Stratum 4 froze almost unworn, which is the point of the layer, but it
    left roots identical to their English sources: nether, warden, trial.
    This is the one change the layer does make, and it is a natural one, a
    non-rhotic reduction of unstressed final /er/ to a plain vowel.

    Nether gives netha. Guarded on syllable count, so a monosyllable whose
    ONLY nucleus is that e is untouched: per stays per.
    """
    segs = tokenize(word)
    if len(segs) >= 2 and segs[-1] == "r" and segs[-2] == "e" and nuclei(word) > 1:
        return untokenize(segs[:-2]) + "a"
    return word


def apply_p9(word, floor):
    """
    P9. MINIMAL WORD. A phonotactic REPAIR, not a sound change: it fires
    because the output is too small to be a word, not because history did
    something to it. Counted in SEGMENTS, so chest (4 segments) and shard
    (4 segments) are treated alike where the letter count disagreed.
    """
    if seg_len(word) < floor and word and not is_vowel(tokenize(word)[-1]):
        return apply_s6(word)
    return word


# ============================================================ PHONOTACTICS
LEGAL_CODA_2 = {"rp", "rt", "rd", "rl", "rn", "rm", "lt", "ld", "lk", "lm",
                "ng", "nt", "nd", "mb", "mf", "st", "sk", "kt", "nth", "tch",
                "sh", "th", "ch", "lf", "rk", "rf", "nk", "mp", "sp", "ft",
                "wl", "wr"}


def legal_onset(onset):
    """
    Onsets are judged by SONORITY, not by an enumerated list.
    An enumerated list built from 175 roots only contains clusters that
    happened to appear, so it rejected sn-, kr-, pr- and other ordinary
    onsets. Two shapes are legal:
      obstruent + liquid/glide   (bl, br, kr, pr, fl, thr...)
      s + consonant              (sk, sm, sn, sp, st, str, spr...)
    plus the digraphs, which are single segments.
    """
    segs = tokenize(onset)
    if len(segs) == 1:
        return True
    if segs[0] in DIGRAPH_ONSETS:
        return len(segs) == 2 and segs[1] in LIQUIDS_GLIDES
    if segs[0] == "s":
        rest = segs[1:]
        if len(rest) == 1:
            return (rest[0] in OBSTRUENTS or rest[0] in NASALS
                    or rest[0] in LIQUIDS_GLIDES or rest[0] in DIGRAPH_ONSETS)
        # s + digraph, or s + obstruent + liquid. The shift produces the
        # first legitimately: Stonecutter > sthon, Storm > sthur.
        if len(rest) == 2:
            return rest[0] in OBSTRUENTS and rest[1] in LIQUIDS_GLIDES
        return False
    if len(segs) == 2:
        # w plus a liquid is legal. Metathesis produces it (Wool > wlu) and
        # the lexicon carries it, so the constraint table has to allow it.
        if segs[0] == "w" and segs[1] in "lr":
            return True
        return segs[0] in OBSTRUENTS and segs[1] in LIQUIDS_GLIDES
    return False


def legal(word, stratum=None, compound=False):
    """
    Check a root against the phonotactic constraints. Returns (ok, reasons).

    STRATA 4 AND 5 ARE EXEMPT from the cluster rules. Stratum 4 froze during
    the collapse almost unworn, so it keeps codas the settled vocabulary would
    never carry (nethr, grindst). Stratum 5 is post-collapse and undergoes no
    sound change at all. Minimal wear is the point of both, so treating their
    shapes as violations was fighting the design. Distinctness and the vowel
    requirement still apply everywhere.
    """
    problems = []
    if not word:
        return False, ["empty"]
    segs = tokenize(word)
    if not any(is_vowel(s) for s in segs):
        problems.append("no vowel")
    # P8 DISTINCTNESS: a root cannot be one segment repeated. Guarded on
    # length, because a single long nucleus is one segment and Eye gives
    # exactly that: ai is a word, aa is not.
    if len(segs) > 1 and len(set(segs)) == 1:
        problems.append(f"single repeated segment '{word}'")

    allowed = set(CONSONANT_SEGMENTS)
    if str(stratum) in ("4", "5"):
        allowed |= LATE_ONLY_SEGMENTS
    outside = [s for s in segs
               if not (is_vowel(s) or s in allowed or s in DIGRAPHS)]
    if outside:
        problems.append(f"not in the inventory {sorted(set(outside))}")

    if str(stratum).lower() in ("4", "5"):
        return (not problems), problems
    if compound:
        # A compound is bracketed: each half was already checked as a root,
        # and the seam between them is not an onset or a coda of anything.
        return (not problems), problems

    onset = onset_of(word)
    if len(onset) > 3:
        problems.append(f"onset too long '{untokenize(onset)}'")
    elif len(onset) > 1 and not legal_onset(untokenize(onset)):
        problems.append(f"illegal onset '{untokenize(onset)}'")

    coda = coda_of(word)
    if len(coda) == 2 and untokenize(coda) not in LEGAL_CODA_2:
        problems.append(f"illegal coda cluster '{untokenize(coda)}'")
    if len(coda) > 2:
        problems.append(f"coda too long '{untokenize(coda)}'")

    # Count with nuclei(), not syllabify(). syllabify() splits a vowel
    # sequence because S1b needs that split to produce dia, so measuring the
    # ceiling against it failed every root with a diphthong: aira parsed as
    # a-ir-a and was reported illegal.
    #
    # The ceiling is THREE, loosened from two, and it applies to SIMPLE
    # roots in the settled strata only. It describes what a single borrowed
    # word looks like after the sound changes wore it down. It says nothing
    # about a compound, which joins two roots that each already satisfy it,
    # and nothing about strata 4 and 5, which froze before the wear that
    # produces short roots in the first place. Applying it to those was
    # measuring them against a process they never went through.
    if not compound and nuclei(word) > MAX_NUCLEI:
        problems.append(f"more than {MAX_NUCLEI} syllables")

    return (not problems), problems


# ============================================== S10: STRATUM 2 SHAPE
# Measured across all 63 stratum-2 roots: the surviving SOURCE material is
# always exactly one syllable. Zero exceptions. The 41 that look longer are
# 35 root-plus-suffix, 5 diphthongs miscounted by syllabify(), and 1 compound.

SUFFIXES = ("a", "i", "o")


def strip_suffix(word):
    """
    S10c SUFFIX EXEMPTION. Remove a final -a, -i or -o if a core remains.
    Guarded: the core must keep a vowel of its own. Otherwise klo, snwo and
    spwa get stripped to kl, snw, spw: vowel-less strings that are not
    cores at all. In those roots the final vowel IS the nucleus.
    """
    segs = tokenize(word)
    if (len(segs) > 2 and segs[-1] in SUFFIXES and not is_vowel(segs[-2])
            and any(is_vowel(s) for s in segs[:-1])):
        return untokenize(segs[:-1])
    return word


def stratum2_shape(word, is_compound=False):
    """
    S10a MONOSYLLABIC CORE. A stratum-2 root's source material is exactly one
    syllable. Compounds are exempt: they join two roots that each satisfy it.
    Returns (ok, reason).
    """
    if is_compound:
        return True, "compound, exempt"
    core = strip_suffix(word)
    n = nuclei(core)
    if n == 1:
        return True, ""
    return False, f"core '{core}' has {n} nuclei; stratum 2 allows exactly 1"


# ================================================================ STRATA
# First-class, so the history stays visible. Revision 1 translated 2b into
# pipeline stratum 3 and carried the difference in a side flag.
#
#   s1_keep      S1 does not fire at or below this many syllables. Stratum 4
#                froze almost unworn, so a word of two syllables or fewer
#                arrives whole: Orange gives orang, not or.
#   shift        the S8 chain shift applies
#   s10a         the monosyllabic-core shape rule is enforced
#   wear         the monosyllabic wear rules S3b and S3d apply
#   floor        minimal-word floor in SEGMENTS, 0 disables the repair
#   sound_change any historical rule at all applies
#
# 2b exists because sound changes are time-bounded events. A word borrowed
# after the shift stopped operating never undergoes it, however core it
# later becomes: English shirt (native, shifted) beside skirt (Norse,
# borrowed after).

STRATA = {
    "2a": {"shift": True,  "s10a": True,  "wear": True,  "floor": 3, "sound_change": True, "s1_keep": 0},
    "2b": {"shift": False, "s10a": True,  "wear": True,  "floor": 3, "sound_change": True, "s1_keep": 0},
    "3":  {"shift": False, "s10a": False, "wear": True,  "floor": 3, "sound_change": True, "s1_keep": 0},
    "4":  {"shift": False, "s10a": False, "wear": False, "floor": 0, "sound_change": False, "s1_keep": 2},
    "5":  {"shift": False, "s10a": False, "wear": False, "floor": 0, "sound_change": False, "s1_keep": 0},
}


def norm_stratum(s):
    """Accept 2, 2a, 2b, 3, 4, 5 in any casing, as text or number."""
    t = str(s).strip().lower()
    if t in ("2", "2a", "2.0"):
        return "2a"
    if t == "2b":
        return "2b"
    if t.startswith("3"):
        return "3"
    if t.startswith("4"):
        return "4"
    if t.startswith("5"):
        return "5"
    return "3"


# ========================================================== RULE PIPELINE
# Ordering is DATA. The comments below record why each pair is ordered the
# way it is, and the pipeline list is the only place that order lives.

HISTORICAL, SPORADIC, REPAIR, SYNCHRONIC = "historical", "sporadic", "repair", "synchronic"


class Rule:
    """One step in a stratum pipeline."""

    def __init__(self, code, name, category, func, needs=None):
        self.code = code
        self.name = name
        self.category = category
        self.func = func
        self.needs = needs or (lambda ctx: True)

    def __repr__(self):
        return f"<{self.code} {self.category}>"


# LOSSES BEFORE SHIFTS: S5 deletes an initial consonant, so it has to act on
# the unshifted form. Running S8 first would delete a segment the shift had
# already changed.
R_S5 = Rule("S5", "aphaeresis", SPORADIC,
            lambda w, ctx: apply_s5(w),
            lambda ctx: ctx["aphaeresis"])

# S8 BLEEDS the old S4: the shift turns a final d into t, which is no longer
# a voiced stop, so final-stop loss finds nothing to delete. Wood > wut.
R_S8 = Rule("S8", "the shift", HISTORICAL,
            lambda w, ctx: apply_s8(w),
            lambda ctx: STRATA[ctx["stratum"]]["shift"])

# S7 BEFORE PADDING. Breaking changes the STEM vowel; S6 and P9 add material
# after it. Running the pad first gave Iron > ir > iri > airi, not ir > air.
R_S7 = Rule("S7", "breaking", SPORADIC,
            lambda w, ctx: apply_s7(w),
            lambda ctx: ctx["breaking"])

# ADDITIONS BEFORE REDUCTIONS: S6 must precede S3, or a cluster that would be
# legal once the -a arrives gets reduced first. anv > anva, not anv > an > ana.
R_S6 = Rule("S6", "paragoge", SPORADIC,
            lambda w, ctx: apply_s6(w, ctx["pad_vowel"]),
            lambda ctx: ctx["paragoge"])

R_P9 = Rule("P9", "minimal word", REPAIR,
            lambda w, ctx: apply_p9(w, STRATA[ctx["stratum"]]["floor"]),
            lambda ctx: ctx["minimal_word"] and not ctx["paragoge"])

R_S3 = Rule("S3", "cluster reduction", HISTORICAL,
            lambda w, ctx: apply_s3(w))

R_S3B = Rule("S3b", "monosyllabic coda", HISTORICAL,
             lambda w, ctx: apply_s3b(w),
             lambda ctx: ctx["mono_source"] and STRATA[ctx["stratum"]]["wear"])

R_S3C = Rule("S3c", "compensatory raising", HISTORICAL,
             lambda w, ctx: apply_s3c(w),
             lambda ctx: ctx["mono_source"] and STRATA[ctx["stratum"]]["wear"])

# S9 runs last: metathesis reorders whatever the other rules produced.
R_P10 = Rule("P10", "minimum content", REPAIR,
             lambda w, ctx: apply_p10(w))

R_S9 = Rule("S9", "metathesis", SPORADIC,
            lambda w, ctx: apply_s9(w, ctx["metathesis_site"]),
            lambda ctx: ctx["metathesis"] is not False and ctx["metathesis"] is not None)

PIPELINE_DEEP = [R_S5, R_S8, R_S7, R_S6, R_P9, R_S3, R_S3B, R_S3C, R_P10, R_S9]
R_S11 = Rule("S11", "final -er to -a", HISTORICAL,
             lambda w, ctx: apply_s11(w),
             lambda ctx: ctx["stratum"] == "4")

# Stratum 4: S0, S1 and the sporadics, plus S11, the one change of its own.
PIPELINE_COLLAPSE = [R_S5, R_S11, R_S6, R_S9]
PIPELINE_FROZEN = []                        # stratum 5: nothing at all

PIPELINES = {"2a": PIPELINE_DEEP, "2b": PIPELINE_DEEP, "3": PIPELINE_DEEP,
             "4": PIPELINE_COLLAPSE, "5": PIPELINE_FROZEN}


# ================================================================ DRIVER
FORCED_TYPE = "Forced"


def is_forced(type_value):
    """
    A FORCED root bypasses every rule. Some words villagers took whole, as a
    label rather than a word: an acronym or a sign they copied without
    parsing. poi from Points of Interest is one: there is no derivation to
    model, so the converter must not try.
    """
    return str(type_value).strip().lower() == "forced"


def convert(source, stratum, metathesis=False, aphaeresis=False,
            paragoge=False, breaking=False, minimal_word=True,
            pad_vowel=None, verbose=True):
    """Derive a root. Returns (root, legal, problems)."""
    strat = norm_stratum(stratum)
    raw = source.lower().strip()
    w = normalise(raw, strat)
    steps = [("source", raw), ("G2P normalised", w)]

    if not STRATA[strat]["sound_change"] and strat == "5":
        # Post-collapse. Villagers could no longer coin or borrow, so anything
        # from this era is built from existing roots and no rule touches it.
        steps.append(("stratum 5: no change", w))
        ok, problems = legal(w, strat)
        if verbose:
            _report(source, strat, steps, w, ok, problems)
        return w, ok, problems

    stress, mode = s0_stress(w)
    steps.append(("S0 stress", f"syllable {stress}, {mode}"))

    ctx = {"stratum": strat,
           "mono_source": len(syllabify(w)) == 1,
           "aphaeresis": aphaeresis, "paragoge": paragoge,
           "breaking": breaking, "metathesis": metathesis,
           "metathesis_site": 0 if metathesis is True else int(metathesis or 0),
           "minimal_word": minimal_word,
           "pad_vowel": pad_vowel}

    keep = STRATA[strat]["s1_keep"]
    if keep and nuclei(w) <= keep:
        steps.append((f"S1 suspended, {nuclei(w)} syllables at stratum {strat}", w))
    else:
        w = apply_s1(w, stress, mode)
        steps.append((f"S1 truncation ({mode})", w))

    for rule in PIPELINES[strat]:
        if rule.needs(ctx):
            before, w = w, rule.func(w, ctx)
            if w != before:
                steps.append((f"{rule.code} {rule.name}", w))

    ok, problems = legal(w, strat)
    if verbose:
        _report(source, strat, steps, w, ok, problems)
    return w, ok, problems


def _report(source, strat, steps, w, ok, problems):
    print(f"\n  {source}  (stratum {strat})")
    print("  " + "-" * 46)
    for name, val in steps:
        print(f"  {name:<24} {val}")
    print("  " + "-" * 46)
    print(f"  RESULT   {w}")
    print(f"  LEGAL    {'yes' if ok else 'NO: ' + '; '.join(problems)}")


# ================================================================ SOLVER
# A solver that can reach any root by SOME combination is not evidence of
# anything. Cost makes the difference visible: a root that falls out with no
# sporadic rules is regular, and a root that needs three rules, a pad vowel
# and a different stratum is a candidate for redesign, not a success.

COSTS = {"sporadic": 1, "minimal_word": 1, "pad_vowel": 2, "stratum": 3}

SPORADIC_FLAGS = [("S5", "aphaeresis"), ("S6", "paragoge"),
                  ("S7", "breaking"), ("S9", "metathesis")]


def _candidates(stratum):
    """
    Every derivation the solver will try for one stratum, cheapest first.
    Each is (cost, label, kwargs).

    Cost is the point of the exercise. A solver that can reach any root by
    SOME combination proves nothing, so the report has to say how much the
    derivation cost:
        exact regular derivation        0
        one sporadic rule               1 each
        minimal word suppressed         1
        non-default paragogic vowel     2
        different historical stratum    3
    A root that only derives at cost 6 is a candidate for redesign, not a
    success.
    """
    import itertools
    out = []
    for n in range(len(SPORADIC_FLAGS) + 1):
        for combo in itertools.combinations(SPORADIC_FLAGS, n):
            label = " ".join(c[0] for c in combo) or "(regular)"
            kw = {c[1]: True for c in combo}
            base_cost = COSTS["sporadic"] * n
            sites = range(2) if kw.get("metathesis") else [0]
            for site in sites:
                for mw in (True, False):
                    pads = ([None] if not (mw or kw.get("paragoge"))
                            else [None, "o", "i", "u", "e"])
                    for pv in pads:
                        cost = base_cost
                        if not mw:
                            cost += COSTS["minimal_word"]
                        if pv is not None:
                            cost += COSTS["pad_vowel"]
                        v = dict(kw, minimal_word=mw, pad_vowel=pv,
                                 stratum=stratum)
                        if kw.get("metathesis"):
                            v["metathesis"] = site
                        out.append((cost, label, v))
    out.sort(key=lambda t: t[0])
    return out


def solve(root, source, stratum, type_value=None):
    """
    Cheapest derivation of a recorded root.
    Returns (cost, label, stratum_used, output). cost None means no
    derivation was found at any price.
    """
    if is_forced(type_value):
        return 0, "FORCED", stratum, root

    recorded = norm_stratum(stratum)
    best = None
    for cost, label, kw in _candidates(recorded):
        kw = dict(kw)
        strat = kw.pop("stratum")
        got, _, _ = convert(source, strat, verbose=False, **kw)
        if got == root:
            best = (cost, label, strat, got)
            break

    if best is None:
        # The recorded stratum cannot produce the root. Another one might,
        # which usually means the recorded stratum is wrong. Anva is the
        # worked case: stratum 2a applies the shift and yields anfa, while
        # stratum 3 yields anva exactly.
        for alt in ("2a", "2b", "3", "4"):
            if alt == recorded:
                continue
            for cost, label, kw in _candidates(alt):
                kw = dict(kw)
                strat = kw.pop("stratum")
                got, _, _ = convert(source, strat, verbose=False, **kw)
                if got == root:
                    cand = (cost + COSTS["stratum"], label, strat, got)
                    if best is None or cand[0] < best[0]:
                        best = cand
                    break
    if best is None:
        got, _, _ = convert(source, recorded, verbose=False)
        return None, None, recorded, got
    return best


# ================================================================ LEXICON
WORKBOOK = "emerald_capitalism_language_system.xlsx"



# ===========================================================================
# WORD FORMATION. Rule C1 compounding and rule A1 affixation.
#
# These used to live only as prose on the Grammar sheet, and build.py carried
# 18 hand-written compounds. Both are now rules, so a change to the phonology
# propagates instead of needing 18 manual edits.
#
# C1  Head-final. A SIMPLE modifier clips to its first syllable when the joined
#     form runs over two syllables. A compound modifier does not clip again.
#     The head never clips.
#
# A1  Insert n between a VOWEL-FINAL stem and a VOWEL-INITIAL suffix. This
#     replaces the old rule where a stem's final vowel dropped before -i, which
#     needed a guard for monosyllables. A1 never deletes, so the guard is gone.
#     -na needs no repair, being consonant-initial already. A1 is the rule that
#     choice anticipated.
# ===========================================================================

# Compounds formed before C1 became productive. Stored whole, not rebuilt.
LEXICALIZED = {"muknek", "ookareg", "ruporta"}

STOPS = {"p", "b", "t", "d", "k", "g"}
MAX_SEAM = 3
VOWEL_INITIAL_SUFFIXES = {"a", "i", "o", "in", "ek", "ur"}


def _coda_ok(coda):
    if len(coda) <= 1:
        return True
    if len(coda) == 2:
        return "".join(coda) in LEGAL_CODA_2
    return False


def clip(modifier):
    """Reduce to the first syllable. syllabify() is maximal-coda, so vila
    gives vil and emra gives emr."""
    parts = syllabify(modifier)
    return parts[0] if parts else modifier


def junction(mod, head, stratum="5"):
    """Repair the seam between a modifier and a head."""
    steps = []
    mseg, hseg = tokenize(mod), tokenize(head)

    # Vowel before vowel drops the first. Grammar sheet, substrate junction.
    if mseg and hseg and is_vowel(mseg[-1]) and is_vowel(hseg[0]):
        steps.append(f"junction: vowel before vowel, dropped /{mseg[-1]}/")
        mseg = mseg[:-1]

    # Degemination across the seam. kov + volt gives kovolt.
    if mseg and hseg and mseg[-1] == hseg[0] and not is_vowel(mseg[-1]):
        steps.append(f"degemination across the seam, dropped /{mseg[-1]}/")
        mseg = mseg[:-1]

    # Stratum 5 seam lenition. A STOP deletes before a consonant. enk + boon
    # gives enboon. Restricted to true stops: extending it to all obstruents
    # destroys hahlapa and the whole emhah set.
    if (str(stratum).strip() == "5" and mseg and hseg
            and mseg[-1] in STOPS and not is_vowel(hseg[0])):
        steps.append(f"stratum 5 seam lenition, dropped /{mseg[-1]}/")
        mseg = mseg[:-1]

    # Trim the modifier coda while it is illegal or the seam is overlong.
    onset = onset_of(untokenize(hseg))
    while mseg:
        coda = coda_of(untokenize(mseg))
        if not coda:
            break
        if _coda_ok(coda) and len(coda) + len(onset) <= MAX_SEAM:
            break
        steps.append(f"seam trim, dropped /{coda[-1]}/")
        mseg = mseg[:-1]

    return untokenize(mseg + hseg), steps


def affix(stem, sfx):
    """Rule A1. Returns (form, steps)."""
    sfx = sfx.lstrip("-")
    seg = tokenize(stem)
    if sfx in VOWEL_INITIAL_SUFFIXES and seg and is_vowel(seg[-1]):
        return stem + "n" + sfx, [f"A1 hiatus repair: n inserted before -{sfx}"]
    return stem + sfx, ["affix, junction rule suspended"]


def compound(modifier, head, modifier_is_compound=False, stratum="5"):
    """Rule C1. Returns (form, steps). Pass a head of '-x' for affixation."""
    if head.startswith("-"):
        return affix(modifier, head)

    steps = []
    if modifier_is_compound:
        steps.append("compound modifier, no reclip")
    elif nuclei(modifier + head) > 2:
        c = clip(modifier)
        if c != modifier:
            steps.append(f"C1 clip: {modifier} -> {c}")
        modifier = c

    form, jsteps = junction(modifier, head, stratum)
    return form, steps + jsteps


def build_from_parts(parts, known_compounds=(), stratum="5"):
    """'emhah + die' or 'kool + -a' -> the built form."""
    mod, head = [p.strip() for p in parts.split("+", 1)]
    return compound(mod, head,
                    modifier_is_compound=mod in set(known_compounds),
                    stratum=stratum)[0]

def load_lexicon(path=WORKBOOK):
    """
    Read the Alphabetical Index into entries. The sporadic rules a root took
    are LEXICAL DATA and already live in the workbook, so the converter reads
    them instead of asking for command-line flags.
    """
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True)
    ws = wb["Alphabetical Index"]
    header, entries = None, []
    for row in ws.iter_rows(values_only=True):
        if row and row[0] == "Root":
            header = list(row)
            continue
        if header and row and row[0]:
            d = {header[i]: row[i] for i in range(len(header))}
            entries.append({
                "root": str(d["Root"]).strip(),
                "source": str(d["Derivation"] or "").strip(),
                "stratum": norm_stratum(d["Stratum"]),
                "type": str(d["Type"] or "").strip(),
                "sporadic": [c for c in ("S5", "S6", "S7", "S9")
                             if c in str(d["Sporadic rules"] or "")],
                "meaning": d["Meaning"],
            })
    return entries


def metathesis_site(sporadic):
    """
    S9 can fire at more than one site and which one moved is a property of the
    word, so the lexicon records it as S9(1). Bare S9 means site 0.
    Wool has two sites: site 0 gives owl, which is an English word, and site 1
    gives wlo.
    """
    import re as _re
    m = _re.search(r"S9\((\d+)\)", " ".join(sporadic) if isinstance(sporadic, list) else str(sporadic))
    return int(m.group(1)) if m else 0


def entry_kwargs(entry):
    return {"aphaeresis": "S5" in entry["sporadic"],
            "paragoge": "S6" in entry["sporadic"],
            "breaking": "S7" in entry["sporadic"],
            "metathesis": (metathesis_site(entry.get("sporadic_raw", entry["sporadic"]))
                           if "S9" in entry["sporadic"] else False)}


def audit(path=WORKBOOK):
    """Check every recorded root against the constraints, at its own stratum."""
    entries = load_lexicon(path)
    print(f"\n  Auditing {len(entries)} roots\n  " + "-" * 52)
    bad = 0
    for e in entries:
        is_compound = ("+" in e["source"]
                       or e["type"].lower().startswith("compound"))
        ok, problems = legal(e["root"], e["stratum"], compound=is_compound)
        if e["stratum"] in ("2a", "2b") and ok:
            ok2, why = stratum2_shape(e["root"],
                                      e["type"].lower().startswith("compound"))
            if not ok2:
                ok, problems = False, [why]
        if not ok:
            bad += 1
            print(f"  {e['root']:<10} str {e['stratum']:<3} {'; '.join(problems)}")
    print("  " + "-" * 52)
    print(f"  {bad} of {len(entries)} violate the constraints\n")


def regenerate(path=WORKBOOK, show_misses=True):
    """
    Derive every borrowed root from its recorded source, stratum and sporadic
    rules, then report exact reproduction and the cost of the rest.
    """
    entries = load_lexicon(path)
    exact, misses, costs = 0, [], {}
    testable = [e for e in entries
                if e["type"].lower() == "borrowed" and e["source"]
                and " " not in e["source"]]
    for e in testable:
        got, ok, _ = convert(e["source"], e["stratum"], verbose=False,
                             **entry_kwargs(e))
        if got == e["root"]:
            exact += 1
            costs[0] = costs.get(0, 0) + 1
        else:
            cost, label, strat, out = solve(e["root"], e["source"], e["stratum"])
            key = cost if cost is not None else "none"
            costs[key] = costs.get(key, 0) + 1
            misses.append((e["root"], e["source"], e["stratum"], got, cost, label, strat))
    print(f"\n  REGENERATION over {len(testable)} borrowed roots")
    print("  " + "-" * 52)
    print(f"  exact from recorded data   {exact}  ({100 * exact / len(testable):.0f} percent)")
    for k in sorted(costs, key=lambda x: (x == "none", x)):
        tag = "no derivation at any price" if k == "none" else f"cheapest derivation cost {k}"
        print(f"  {tag:<28} {costs[k]}")
    if show_misses:
        print("  " + "-" * 52)
        print(f"  {'root':<10}{'source':<14}{'str':<5}{'got':<10}{'cost':<6}{'via'}")
        for root, src, strat, got, cost, label, used in misses:
            c = "--" if cost is None else str(cost)
            via = "" if cost is None else f"{label} @ {used}"
            print(f"  {root:<10}{src:<14}{strat:<5}{got:<10}{c:<6}{via}")
    print()


# ================================================================ CLI
if __name__ == "__main__":
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    flags = {a for a in sys.argv[1:] if a.startswith("--")}

    try:
        if "--audit" in flags:
            audit()
        elif "--regenerate" in flags:
            regenerate(show_misses="--quiet" not in flags)
        elif "--solve" in flags and len(args) >= 3:
            cost, label, strat, out = solve(args[0], args[1], args[2])
            if cost is None:
                print(f"  no derivation of {args[0]}; regular output is {out}")
            else:
                print(f"  {args[0]} = {args[1]} @ stratum {strat} via {label}, cost {cost}")
        elif len(args) >= 2:
            convert(args[0], args[1],
                    metathesis="--metathesis" in flags,
                    aphaeresis="--aphaeresis" in flags,
                    paragoge="--paragoge" in flags,
                    breaking="--breaking" in flags)
        else:
            print(__doc__)
    except FileNotFoundError:
        print(f"  Put {WORKBOOK} in this folder first.")
