#!/usr/bin/env python3
"""Export the substrate naming system from the language workbook.

The workbook's ``Substrate Names`` sheet is a regression oracle, not a runtime
name list.  This converter exports the closed element sets, biome pools,
phonological rules, drift rules, assignment tuning, age suffixes, and
profession bynames, plus a small manually maintained special-name section.
Java assembles the names from that system at runtime.

Usage::

    python3 scripts/convert_villager_names.py
    python3 scripts/convert_villager_names.py --input path/to/book.xlsx
    python3 scripts/convert_villager_names.py --validate-only
"""

from __future__ import annotations

import argparse
import json
import re
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_INPUT = ROOT / "docs" / "lore" / "emerald_capitalism_language_system.xlsx"
DEFAULT_OUTPUT = (ROOT / "src" / "main" / "resources" / "data"
                  / "emeraldcapitalism" / "village_naming"
                  / "villager_names.json")

BIOME_ORDER = ("plains", "savanna", "desert", "taiga", "snowy")
STOPS = "bdgkpt"
VOWELS = "aeiou"

# These are tuning values, not language design.  They are deliberately
# exported so the runtime and the source-data report use the same model.
DRIFT_THRESHOLD = 0.78
DRIFT_FEATURE_SIZES = [4096, 5120, 3072, 6144, 2048, 4096,
                       3072, 5120, 4096, 4096, 2048, 3072]

DRIFT_OPERATIONS = {
    "D1": "drop_final_nasal",
    "D2": "lenite_medial_s_to_h",
    "D3": "lower_i_to_e",
    "D4": "lower_u_to_o",
    "D5": "simplify_geminates",
    "D6": "drop_final_liquid",
    "D7": "change_final_stop_to_nasal",
    "D8": "lenite_medial_k_to_h",
    "D9": "voice_initial_stop",
    "D10": "devoice_initial_stop",
    "D11": "lenite_medial_t_to_s",
    "D12": "lenite_medial_m_to_w",
}

# Manual exceptions. These are deliberately outside the workbook's canonical
# substrate pools and are copied into the generated resource unchanged.
SPECIAL_FIRST_NAMES = ["Kinniken"]
SPECIAL_FIRST_NAME_SELECTION_CHANCE = 0.0001  # 1 in 10,000 assignments


def text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def rows(sheet):
    return [tuple(text(value) for value in row)
            for row in sheet.iter_rows(values_only=True)]


def lookup(rule_rows, area: str, item: str) -> str:
    for row in rule_rows:
        if len(row) >= 3 and row[0] == area and row[1] == item:
            return row[2]
    raise ValueError(f"Missing workbook rule: {area} / {item}")


def parse_list(value: str) -> list[str]:
    result = []
    for part in value.split(","):
        part = part.strip()
        if not part:
            continue
        # The rare-element row has a prose note after the data.
        result.append(part.split(".", 1)[0].strip())
    return result


def simplify_geminates(word: str) -> str:
    output: list[str] = []
    for char in word:
        if output and output[-1] == char and char not in VOWELS:
            continue
        output.append(char)
    return "".join(output)


def junction(first: str, second: str) -> str:
    if first[-1] in STOPS and second[0] not in VOWELS:
        # This order is important: gemination wins over stop deletion.
        if first[-1] == second[0]:
            return first + second
        return first[:-1] + second
    return first + second


def biome_form(base: str, biome: str) -> str:
    if biome == "desert":
        return base.replace("e", "i")
    if biome == "savanna":
        return simplify_geminates(base)
    if biome == "taiga":
        return base.replace("o", "u")
    if biome == "snowy" and base[-1] in STOPS:
        return base[:-1]
    return base


def parse_biome_rules(rule_rows) -> dict[str, dict[str, str]]:
    rules: dict[str, dict[str, str]] = {
        "plains": {"operation": "none"},
        "savanna": {"operation": "simplify_geminates"},
        "desert": {"operation": "replace", "from": "e", "to": "i"},
        "taiga": {"operation": "replace", "from": "o", "to": "u"},
        "snowy": {"operation": "delete_final_stops", "segments": STOPS},
    }
    for biome, rule in ((row[0][6:].lower(), row[2]) for row in rule_rows
                        if row[0].startswith("BIOME ")):
        if biome not in rules:
            raise ValueError(f"Unknown biome in workbook: {biome}")
        # Keep the workbook text auditable without allowing a typo to silently
        # change the runtime interpretation.
        if biome == "desert" and "e > i" not in rule:
            raise ValueError("Desert biome rule no longer says e > i")
        if biome == "savanna" and "geminates simplify" not in rule:
            raise ValueError("Savanna biome rule no longer simplifies geminates")
        if biome == "taiga" and "o > u" not in rule:
            raise ValueError("Taiga biome rule no longer says o > u")
        if biome == "snowy" and "final STOPS delete" not in rule:
            raise ValueError("Snowy biome rule no longer deletes final stops")
    return rules


def parse_profession_key(item: str) -> str:
    base = item.split(" (", 1)[0]
    return re.sub(r"[^a-z0-9]", "", base.lower())


def remove_agent_suffix(byname: str) -> str:
    """Recover the regional profession stem from its generated ``-i`` form.

    A1 inserts ``n`` before ``-i`` after a vowel-final stem.  The workbook's
    byname columns are the checked output of that rule, so stripping ``ni``
    before the plain ``i`` preserves the regional stem for the runtime to
    rebuild and verify with the same rule.
    """
    if byname.endswith("ni"):
        return byname[:-2]
    if byname.endswith("i"):
        return byname[:-1]
    return byname


def build_system(workbook) -> dict[str, Any]:
    substrate_rows = rows(workbook["Substrate"])
    first = parse_list(lookup(substrate_rows, "FIRST ELEMENTS", "All 20"))
    second_common = parse_list(lookup(substrate_rows, "SECOND ELEMENTS", "Common 16"))
    second_rare = parse_list(lookup(substrate_rows, "SECOND ELEMENTS", "Rare 4"))

    pools: dict[str, dict[str, list[str]]] = {}
    for biome in BIOME_ORDER:
        area = f"POOL {biome.title()}"
        pools[biome] = {
            "first": parse_list(lookup(substrate_rows, area, next(
                item for row_area, item, _ in substrate_rows
                if row_area == area and item.startswith("First ")))),
            "second": parse_list(lookup(substrate_rows, area, next(
                item for row_area, item, _ in substrate_rows
                if row_area == area and item.startswith("Second ")))),
        }

    drift_rules = []
    for area, item, rule in substrate_rows:
        if not area.startswith("DRIFT D"):
            continue
        rule_id = area.split()[1]
        if rule_id not in DRIFT_OPERATIONS:
            raise ValueError(f"Unknown drift rule in workbook: {rule_id}")
        drift_rules.append({
            "id": rule_id,
            "name": item,
            "description": rule,
            "operation": DRIFT_OPERATIONS[rule_id],
        })
    drift_rules.sort(key=lambda entry: int(entry["id"][1:]))
    if [entry["id"] for entry in drift_rules] != [f"D{i}" for i in range(1, 13)]:
        raise ValueError("Substrate must contain exactly D1 through D12")

    profession_rows = rows(workbook["Names and Professions"])
    header_index = next(i for i, row in enumerate(profession_rows)
                        if row[:3] == ("Type", "Item", "Root"))
    professions: dict[str, dict[str, Any]] = {}
    for row in profession_rows[header_index + 1:]:
        if not row or row[0] != "PROFESSION":
            continue
        item, root = row[1], row[2]
        bynames = {}
        for biome, column in (("plains", 3), ("desert", 4), ("savanna", 5),
                              ("taiga", 6), ("snowy", 7)):
            if len(row) > column and row[column]:
                bynames[biome] = row[column].lower()
        regional_roots = {
            biome: remove_agent_suffix(value)
            for biome, value in bynames.items()
        }
        professions[parse_profession_key(item)] = {
            "label": item,
            "root": root or None,
            "bynames": bynames,
            "regional_roots": regional_roots,
            "note": row[8] if len(row) > 8 else "",
        }

    return {
        "format": 1,
        "source": {
            "workbook": "emerald_capitalism_language_system.xlsx",
            "sheets": ["Substrate", "Substrate Names", "Names and Professions"],
            "substitute_for": "Substrate Names is a verification oracle, not a shipped list.",
        },
        "elements": {
            "first": first,
            "second": second_common + second_rare,
            "second_frequency": {
                "common": second_common,
                "rare": second_rare,
                "rare_weight": 0.05,
            },
        },
        "pools": pools,
        "junction": {
            "stop_segments": STOPS,
            "vowels": VOWELS,
            "stop_deletes_before_consonant": True,
            "identical_consonants_geminate": True,
            "gemination_wins_over_stop_deletion": True,
        },
        "biome_rules": parse_biome_rules(substrate_rows),
        "drift_assignment": {
            "model": "independent_low_frequency_value_noise",
            "threshold": DRIFT_THRESHOLD,
            "feature_sizes": DRIFT_FEATURE_SIZES,
            "min_rules": 1,
            "max_rules": 2,
            "inverse_groups": [["D9", "D10"]],
            "rule_order": "ascending_rule_index",
            "shared_across_biomes": True,
        },
        "drift_rules": drift_rules,
        "age_suffixes": {"child": "ek", "adult": "in", "elder": "ur"},
        "nitwit_suffix": "ek",
        "origin_particle": {"prefix": "Se", "enabled": True,
                             "note": "Rendered from the persisted origin village once its name is determined."},
        "mayor_origin_particle": {"prefix": "Bi-", "enabled": True,
                                   "note": "Mayor-only prefix for the persisted origin village; do not also render the generic origin particle."},
        "special_first_names": {
            "names": SPECIAL_FIRST_NAMES,
            "selection_chance": SPECIAL_FIRST_NAME_SELECTION_CHANCE,
            "note": "Manual exceptions replace only slot 1; age, profession, and origin remain derived.",
        },
        "professions": professions,
    }


def verify_substrate_names(system: dict[str, Any], workbook) -> dict[str, Any]:
    name_rows = [row for row in rows(workbook["Substrate Names"])[3:]
                 if row[0]]
    counts = Counter(row[0] for row in name_rows)
    expected_counts = {"Plains": 400, "Savanna": 289, "Desert": 169,
                       "Taiga": 100, "Snowy": 81}
    if len(name_rows) != 1039 or dict(counts) != expected_counts:
        raise ValueError(
            f"Substrate Names count mismatch: rows={len(name_rows)} counts={dict(counts)}")

    pools = system["pools"]
    generated = []
    for row in name_rows:
        workbook_biome, local, expected_base, frequency = row[:4]
        biome = workbook_biome.lower()
        pool = pools[biome]
        matching_pairs = [
            (first, second)
            for first in pool["first"]
            for second in pool["second"]
            if junction(first, second).capitalize() == expected_base
        ]
        if len(matching_pairs) != 1:
            raise ValueError(
                f"Could not resolve unique pair for {workbook_biome}/{local}/{expected_base}")
        first, second = matching_pairs[0]
        base = junction(first, second)
        actual_local = biome_form(base, biome).capitalize()
        if base.capitalize() != expected_base or actual_local != local:
            raise ValueError(
                f"Substrate Names mismatch: {workbook_biome}/{local}/{expected_base}; "
                f"got {actual_local}/{base.capitalize()} from {first}+{second}")
        generated.append((workbook_biome, actual_local, base.capitalize(), frequency))

    if generated != name_rows:
        for index, (actual, expected) in enumerate(zip(generated, name_rows), start=4):
            if actual != expected:
                raise ValueError(f"Substrate Names row {index} differs: {actual} != {expected}")
        raise ValueError("Substrate Names row count differs after reproduction")
    return {"rows": len(name_rows), "counts": dict(counts)}


def parse_args():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", default=DEFAULT_INPUT, type=Path,
                        help="source workbook")
    parser.add_argument("--output", default=DEFAULT_OUTPUT, type=Path,
                        help="generated runtime JSON")
    parser.add_argument("--validate-only", action="store_true",
                        help="verify the workbook without writing JSON")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    workbook = load_workbook(args.input, data_only=True, read_only=True)
    system = build_system(workbook)
    verification = verify_substrate_names(system, workbook)
    print(f"  reproduced Substrate Names rows  {verification['rows']}")
    for biome in BIOME_ORDER:
        print(f"    {biome:<7} {verification['counts'][biome.title()]:>3}")
    print(f"  exported drift rules             {len(system['drift_rules'])}")
    print(f"  exported professions              {len(system['professions'])}")
    if args.validate_only:
        return
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(system, indent=2, ensure_ascii=False) + "\n",
                           encoding="utf-8")
    print(f"  written to {args.output}")


if __name__ == "__main__":
    main()
